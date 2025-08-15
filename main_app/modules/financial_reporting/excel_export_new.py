"""
Excel Export Functionality - Recreated to match reference implementation exactly
Creates formatted Excel workbook with all financial reporting sections
"""

from io import BytesIO
import pandas as pd
from datetime import datetime, timedelta
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from .tb_generator import get_income_expense_changes, calculate_nav_changes, generate_trial_balance_from_gl, categorize_account
from .data_processor import get_previous_period_date, format_currency
from ...s3_utils import load_COA_file, load_LP_commitments_file


# Define Excel styles (exactly matching reference implementation)
font_name = "Inter"
font_black = Font(name=font_name, color="000000")
font_color = Font(name=font_name, color="0047BA")
font_title = Font(name=font_name, bold=True, size=20, color="FFFFFF")
font_subtitle = Font(name=font_name, bold=False, size=12, color="FFFFFF")
font_white = Font(name=font_name, color="FFFFFF")
font_black_bold = Font(name=font_name, color="FF000000", bold=True)

alt_fill = PatternFill("solid", fgColor="F2F2F2")
color_fill = PatternFill("solid", fgColor="0047BA")
light_color_fill = PatternFill("solid", fgColor="e1edff")
no_fill = PatternFill()
white_fill = PatternFill("solid", fgColor="FFFFFF")

left = Alignment(horizontal="left")
right = Alignment(horizontal="right")
center = Alignment(horizontal="center")

thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))
top_border = Border(top=Side(style="thin", color="CCCCCC"))


def get_number_format(currency: str = "ETH"):
    """Get number format based on currency"""
    if currency.upper() == "ETH":
        return "#,##0.000000_);(#,##0.000000)"
    elif currency.upper() == "USD":
        return "#,##0.00_);(#,##0.00)"
    else:
        return "#,##0.0000_);(#,##0.0000)"


def get_display_round(currency: str = "ETH"):
    """Get display rounding based on currency"""
    if currency.upper() == "ETH":
        return 6
    elif currency.upper() == "USD":
        return 2
    else:
        return 4


def create_account_statement_sheet(wb, gl_df: pd.DataFrame, fund_name: str, 
                                 report_date: datetime, currency: str = "ETH"):
    """Create Account Statement sheet exactly like reference implementation"""
    ws = wb.create_sheet(title="Account Statement")
    number_format_choice = get_number_format(currency)
    
    # Header formatting (exactly like reference)
    start_of_period = report_date.replace(day=1)
    for row in range(1, 5):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = color_fill
            cell.font = font_subtitle if row > 2 else font_title

    ws["A1"] = fund_name
    ws["A1"].alignment = left
    ws["A1"].font = font_title

    ws["A2"] = "Account statement (unaudited)"
    ws["A2"].alignment = left
    ws["A2"].font = font_title

    ws["A3"] = f"For the period ended {report_date.strftime('%B %d, %Y')}"
    ws["A3"].alignment = left
    ws["A3"].font = Font(name=font_name, color="FFFFFF", size=12, bold=False)

    ws["A4"] = f"Reporting currency: {currency}"
    ws["A4"].alignment = left
    ws["A4"].font = Font(name=font_name, color="FFFFFF", size=12, bold=False)

    ws["E3"] = f"Start of period {start_of_period.strftime('%m/%d/%Y')}"
    ws["E3"].alignment = right
    ws["E3"].font = Font(name=font_name, color="FFFFFF", size=12, bold=False)

    ws["E4"] = f"End of period {report_date.strftime('%m/%d/%Y')}"
    ws["E4"].alignment = right
    ws["E4"].font = Font(name=font_name, color="FFFFFF", size=12, bold=False)

    # Section Title
    ws["A6"] = "Statement of income"
    ws["A6"].font = Font(name=font_name, color="0047BA", bold=True)
    ws["A6"].alignment = left

    # Table Headers
    ws["A7"] = "Income"
    ws["A7"].font = font_black
    ws["A7"].alignment = left

    headers = ["Month to Date", "Quarter to Date", "Year to Date", "Inception to Date"]
    for col_num, header in enumerate(headers, 2):
        cell = ws.cell(row=7, column=col_num, value=header)
        cell.font = font_black
        cell.alignment = right
        cell.border = thin_border
        cell.fill = light_color_fill

    for col in range(1, 6):
        ws.cell(row=7, column=col).fill = light_color_fill

    try:
        # Get income and expense data
        income_df, expense_df = get_income_expense_changes(gl_df, report_date)
        
        # Write Income Rows
        row = 8
        income_totals = [0, 0, 0, 0]
        if not income_df.empty:
            for idx, (_, income_row) in enumerate(income_df.iterrows()):
                fill = alt_fill if idx % 2 == 1 else no_fill
                label = income_row.get('GL_Acct_Name', 'Unknown Income')
                values = [
                    income_row.get('MTD', 0),
                    income_row.get('QTD', 0), 
                    income_row.get('YTD', 0),
                    income_row.get('ITD', 0)
                ]
                
                ws.cell(row=row, column=1, value="  " + label).font = font_black
                ws.cell(row=row, column=1).alignment = left
                ws.cell(row=row, column=1).fill = fill
                
                for j, val in enumerate(values):
                    cell = ws.cell(row=row, column=2 + j, value=val)
                    cell.font = font_black
                    cell.alignment = right
                    cell.fill = fill
                    cell.number_format = number_format_choice
                    income_totals[j] += val
                row += 1

        # Income Total
        ws.cell(row=row, column=1, value="Total income").font = font_black
        for j, val in enumerate(income_totals):
            cell = ws.cell(row=row, column=2 + j, value=val)
            cell.font = font_black
            cell.alignment = right
            cell.border = Border(top=Side(style="thin", color="000000"))
            cell.number_format = number_format_choice
        row += 2

        # Expenses Section
        ws.cell(row=row, column=1, value="Expenses").font = font_black
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = light_color_fill
        row += 1

        expense_totals = [0, 0, 0, 0]
        if not expense_df.empty:
            for idx, (_, expense_row) in enumerate(expense_df.iterrows()):
                fill = alt_fill if idx % 2 == 1 else no_fill
                label = expense_row.get('GL_Acct_Name', 'Unknown Expense')
                values = [
                    expense_row.get('MTD', 0),
                    expense_row.get('QTD', 0),
                    expense_row.get('YTD', 0),
                    expense_row.get('ITD', 0)
                ]
                
                ws.cell(row=row, column=1, value="  " + label).font = font_black
                ws.cell(row=row, column=1).alignment = left
                ws.cell(row=row, column=1).fill = fill
                
                for j, val in enumerate(values):
                    cell = ws.cell(row=row, column=2 + j, value=val)
                    cell.font = font_black
                    cell.alignment = right
                    cell.fill = fill
                    cell.number_format = number_format_choice
                    expense_totals[j] += val
                row += 1

        # Expense Total
        ws.cell(row=row, column=1, value="Total expenses").font = font_black
        for j, val in enumerate(expense_totals):
            cell = ws.cell(row=row, column=2 + j, value=val)
            cell.font = font_black
            cell.alignment = right
            cell.border = Border(top=Side(style="thin", color="000000"))
            cell.number_format = number_format_choice
        row += 2

        # Net Income
        ws.cell(row=row, column=1, value="Net Income").font = font_black_bold
        ws.cell(row=row, column=1).alignment = left
        ws.cell(row=row, column=1).fill = light_color_fill
        for j in range(4):
            net = income_totals[j] - expense_totals[j]
            cell = ws.cell(row=row, column=2 + j, value=net)
            cell.font = font_black_bold
            cell.fill = light_color_fill
            cell.alignment = right
            cell.number_format = number_format_choice

    except Exception as e:
        ws["A8"] = f"Error generating income statement: {str(e)}"
    
    # Add NAV changes to the same sheet
    add_nav_changes_to_sheet(ws, gl_df, report_date, currency)

    # Column Widths
    for col in range(1, 6):
        max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(6, ws.max_row + 1))
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = max_len + 6

    # White Fill Unstyled Cells
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if cell.fill is None or cell.fill == PatternFill():
                cell.fill = white_fill


def add_nav_changes_to_sheet(ws, gl_df: pd.DataFrame, report_date: datetime, currency: str = "ETH"):
    """Add NAV Changes section to existing Account Statement sheet"""
    number_format_choice = get_number_format(currency)
    
    # Find the last used row
    row = ws.max_row + 1
    
    # Add separator
    for col in range(1, 6):
        ws.cell(row=row, column=col).fill = color_fill
    ws.cell(row=row, column=1, value="").font = font_white
    ws.cell(row=row, column=1).alignment = left
    row += 2
    
    # Section title
    ws.cell(row=row, column=1, value="Statement of changes in net asset value").font = Font(name=font_name, color="0047BA", bold=True)
    ws.cell(row=row, column=1).alignment = left
    row += 1

    # Headers (with blank first column like reference - this is different from the export sheet fix)
    headers = ["", "Month to Date", "Quarter to Date", "Year to Date", "Inception to Date"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.font = font_black
        cell.alignment = right
        cell.fill = light_color_fill
        cell.border = thin_border
    row += 1

    try:
        # Get NAV changes data
        nav_df = calculate_nav_changes(gl_df, report_date)
        
        # NAV data rows
        nav_items = ['Beginning balance', 'Capital contributions', 'Distributions', 'Net income (loss)', 'Ending balance']
        black_top_border = Border(top=Side(style="thin", color="000000"))
        
        for i, item in enumerate(nav_items):
            fill = alt_fill if i % 2 == 0 else no_fill
            ws.cell(row=row, column=1).value = item
            ws.cell(row=row, column=1).alignment = left
            ws.cell(row=row, column=1).fill = fill
            ws.cell(row=row, column=1).font = font_black_bold if item == "Ending balance" else font_black

            # Add values for each period
            for j, period in enumerate(['Month to Date', 'Quarter to Date', 'Year to Date', 'Inception to Date']):
                period_row = nav_df[nav_df['Period'] == period] if not nav_df.empty else pd.DataFrame()
                if not period_row.empty:
                    # Map the item names
                    item_map = {
                        'Beginning balance': 'Beginning Balance',
                        'Capital contributions': 'Capital Contributions', 
                        'Distributions': 'Distributions',
                        'Net income (loss)': 'Net Income (Loss)',
                        'Ending balance': 'Ending Balance'
                    }
                    mapped_item = item_map.get(item, item)
                    val = period_row[mapped_item].iloc[0] if mapped_item in period_row.columns else 0
                else:
                    val = 0
                
                cell = ws.cell(row=row, column=2 + j)
                cell.value = val
                cell.font = font_black_bold if item == "Ending balance" else font_black
                cell.alignment = right
                cell.number_format = number_format_choice
                cell.fill = fill
                if item == "Ending balance":
                    cell.border = black_top_border
            row += 1

    except Exception as e:
        ws.cell(row=row, column=1, value=f"Error generating NAV changes: {str(e)}")
        row += 1

    # Footer fill
    for col in range(1, 6):
        ws.cell(row=row, column=col).fill = color_fill


def create_nav_changes_sheet(wb, gl_df: pd.DataFrame, fund_name: str, 
                           report_date: datetime, currency: str = "ETH"):
    """Create NAV Changes sheet"""
    ws = wb.create_sheet(title="NAV Changes")
    number_format_choice = get_number_format(currency)
    
    # Header formatting (exactly like reference)
    start_of_period = report_date.replace(day=1)
    for row in range(1, 5):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = color_fill
            cell.font = font_subtitle if row > 2 else font_title

    ws["A1"] = fund_name
    ws["A1"].alignment = left
    ws["A1"].font = font_title

    ws["A2"] = "Statement of changes in net asset value (unaudited)"
    ws["A2"].alignment = left
    ws["A2"].font = font_title

    ws["A3"] = f"For the period ended {report_date.strftime('%B %d, %Y')}"
    ws["A3"].alignment = left
    ws["A3"].font = Font(name=font_name, color="FFFFFF", size=12, bold=False)

    ws["A4"] = f"Reporting currency: {currency}"
    ws["A4"].alignment = left
    ws["A4"].font = Font(name=font_name, color="FFFFFF", size=12, bold=False)

    ws["E3"] = f"Start of period {start_of_period.strftime('%m/%d/%Y')}"
    ws["E3"].alignment = right
    ws["E3"].font = Font(name=font_name, color="FFFFFF", size=12, bold=False)

    ws["E4"] = f"End of period {report_date.strftime('%m/%d/%Y')}"
    ws["E4"].alignment = right
    ws["E4"].font = Font(name=font_name, color="FFFFFF", size=12, bold=False)
    
    try:
        # Get NAV changes data
        nav_df = calculate_nav_changes(gl_df, report_date)
        
        row = 8
        ws.cell(row=row, column=1, value="Statement of changes in net asset value").font = Font(name=font_name, color="0047BA", bold=True)
        ws.cell(row=row, column=1).alignment = left
        row += 1
        
        # Headers - no blank first column for export sheet (different from account statement)
        headers = ["Month to Date", "Quarter to Date", "Year to Date", "Inception to Date"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            cell.font = font_black
            cell.alignment = right
            cell.fill = light_color_fill
            cell.border = thin_border
        row += 1
        
        # NAV data rows
        nav_items = ['Beginning Balance', 'Capital Contributions', 'Distributions', 'Net Income (Loss)', 'Ending Balance']
        for i, item in enumerate(nav_items):
            fill = alt_fill if i % 2 == 0 else no_fill
            
            # Add values for each period
            for period_idx, period in enumerate(['Month to Date', 'Quarter to Date', 'Year to Date', 'Inception to Date']):
                period_row = nav_df[nav_df['Period'] == period] if not nav_df.empty else pd.DataFrame()
                if not period_row.empty:
                    val = period_row[item].iloc[0] if item in period_row.columns else 0
                else:
                    val = 0
                
                cell = ws.cell(row=row, column=1 + period_idx)
                cell.value = val
                cell.font = font_black_bold if item == "Ending Balance" else font_black
                cell.alignment = right
                cell.number_format = number_format_choice
                cell.fill = fill
                if item == "Ending Balance":
                    cell.border = Border(top=Side(style="thin", color="000000"))
            row += 1

    except Exception as e:
        ws.cell(row=8, column=1, value=f"Error generating NAV changes: {str(e)}")
    
    # Fill unstyled cells
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if cell.fill is None or cell.fill == PatternFill():
                cell.fill = white_fill


def create_trial_balance_sheet(wb, gl_df: pd.DataFrame, fund_name: str, 
                              report_date: datetime, currency: str = "ETH"):
    """Create Trial Balance sheet"""
    ws = wb.create_sheet(title="Trial Balance")
    number_format_choice = get_number_format(currency)
    
    # Calculate comparison date (previous month end)
    from datetime import timedelta
    comp_date = report_date.replace(day=1) - timedelta(days=1)
    
    # Header formatting
    for row in range(1, 5):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.fill = color_fill
            cell.font = font_subtitle if row > 2 else font_title

    ws["A1"] = fund_name
    ws["A1"].alignment = left
    ws["A1"].font = font_title

    ws["A2"] = "Trial balance (unaudited)"
    ws["A2"].alignment = left
    ws["A2"].font = font_title

    ws["A3"] = f"For the period ended {report_date.strftime('%B %d, %Y')}"
    ws["A3"].alignment = left
    ws["A3"].font = Font(name=font_name, color="FFFFFF", size=12, bold=False)

    ws["A4"] = f"Reporting currency: {currency}"
    ws["A4"].alignment = left
    ws["A4"].font = Font(name=font_name, color="FFFFFF", size=12, bold=False)

    ws["F3"] = f"Start of period {comp_date.strftime('%m/%d/%Y')}"
    ws["F3"].alignment = right
    ws["F3"].font = Font(name=font_name, color="FFFFFF", size=12, bold=False)

    ws["F4"] = f"End of period {report_date.strftime('%m/%d/%Y')}"
    ws["F4"].alignment = right
    ws["F4"].font = Font(name=font_name, color="FFFFFF", size=12, bold=False)
    
    try:
        # Generate trial balance data for both dates
        current_tb = generate_trial_balance_from_gl(gl_df, report_date)
        previous_tb = generate_trial_balance_from_gl(gl_df, comp_date)
        
        if current_tb.empty:
            ws["A6"] = "No trial balance data available"
            return
        
        # Merge current and previous balances to calculate changes
        if not previous_tb.empty:
            merged_tb = pd.merge(
                current_tb[['GL_Acct_Number', 'GL_Acct_Name', 'Category', 'Balance']],
                previous_tb[['GL_Acct_Number', 'Balance']],
                on='GL_Acct_Number',
                how='outer',
                suffixes=('_current', '_previous')
            )
        else:
            merged_tb = current_tb.copy()
            merged_tb['Balance_previous'] = 0
            merged_tb['Balance_current'] = merged_tb['Balance']
        
        # Fill missing values
        merged_tb['Balance_current'] = merged_tb['Balance_current'].fillna(0)
        merged_tb['Balance_previous'] = merged_tb['Balance_previous'].fillna(0)
        merged_tb['Change'] = merged_tb['Balance_current'] - merged_tb['Balance_previous']
        
        # Add category info if missing
        if 'Category' not in merged_tb.columns:
            merged_tb['Category'] = merged_tb['GL_Acct_Number'].apply(categorize_account)
        if 'GL_Acct_Name' not in merged_tb.columns:
            from ...s3_utils import load_COA_file
            coa_df = load_COA_file()
            coa_dict = dict(zip(coa_df['GL_Acct_Number'], coa_df['GL_Acct_Name']))
            merged_tb['GL_Acct_Name'] = merged_tb['GL_Acct_Number'].apply(
                lambda x: coa_dict.get(int(x), f"Account {int(x)}")
            )
        
        # Filter out zero balances
        merged_tb = merged_tb[
            (abs(merged_tb['Balance_current']) > 0.000001) |
            (abs(merged_tb['Balance_previous']) > 0.000001) |
            (abs(merged_tb['Change']) > 0.000001)
        ]
        
        # Sort by category then by account number
        category_order = {'Assets': 1, 'Liabilities': 2, 'Capital': 3, 'Income': 4, 'Expenses': 8}
        merged_tb['sort_order'] = merged_tb['Category'].map(category_order).fillna(99)
        merged_tb = merged_tb.sort_values(['sort_order', 'GL_Acct_Number'])
        
        # Section title and headers
        row = 6
        ws.cell(row=row, column=1, value="Trial balance").font = Font(name=font_name, color="0047BA", bold=True)
        ws.cell(row=row, column=1).alignment = left
        row += 1
        
        headers = ["Account category", "GL number", "Account", 
                   report_date.strftime("%m/%d/%Y"), 
                   comp_date.strftime("%m/%d/%Y"), 
                   "Change"]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num, value=header)
            cell.font = Font(name=font_name, bold=True)
            cell.alignment = right if col_num >= 4 else left
            cell.fill = light_color_fill
            cell.border = thin_border
        row += 1
        
        # Data rows
        totals = [0.0, 0.0, 0.0]
        for idx, (_, tb_row) in enumerate(merged_tb.iterrows()):
            fill = alt_fill if idx % 2 == 1 else no_fill
            
            # Category
            ws.cell(row=row, column=1, value=tb_row['Category']).font = font_black
            ws.cell(row=row, column=1).alignment = left
            ws.cell(row=row, column=1).fill = fill
            
            # GL Number  
            ws.cell(row=row, column=2, value=int(tb_row['GL_Acct_Number'])).font = font_black
            ws.cell(row=row, column=2).alignment = left
            ws.cell(row=row, column=2).fill = fill
            
            # Account Name
            ws.cell(row=row, column=3, value=tb_row['GL_Acct_Name']).font = font_black
            ws.cell(row=row, column=3).alignment = left
            ws.cell(row=row, column=3).fill = fill
            
            # Current Balance
            current_val = round(tb_row['Balance_current'], 6)
            cell = ws.cell(row=row, column=4, value=current_val)
            cell.font = font_black
            cell.alignment = right
            cell.fill = fill
            cell.number_format = number_format_choice
            totals[0] += current_val
            
            # Previous Balance
            previous_val = round(tb_row['Balance_previous'], 6)
            cell = ws.cell(row=row, column=5, value=previous_val)
            cell.font = font_black
            cell.alignment = right
            cell.fill = fill
            cell.number_format = number_format_choice
            totals[1] += previous_val
            
            # Change
            change_val = round(tb_row['Change'], 6)
            cell = ws.cell(row=row, column=6, value=change_val)
            cell.font = font_black
            cell.alignment = right
            cell.fill = fill
            cell.number_format = number_format_choice
            totals[2] += change_val
            
            row += 1
        
        # Total row
        ws.cell(row=row, column=1, value="Total").font = Font(name=font_name, bold=True)
        ws.cell(row=row, column=1).alignment = left
        ws.cell(row=row, column=1).fill = light_color_fill
        
        for col_num in range(2, 4):
            ws.cell(row=row, column=col_num).fill = light_color_fill
        
        for col_num, val in enumerate(totals, 4):
            cell = ws.cell(row=row, column=col_num, value=round(val, 6))
            cell.font = Font(name=font_name, bold=True)
            cell.alignment = right
            cell.fill = light_color_fill
            cell.number_format = number_format_choice
            cell.border = Border(top=Side(style="thin", color="000000"))
        
    except Exception as e:
        ws["A6"] = f"Error generating trial balance: {str(e)}"
    
    # Auto-adjust column widths
    for col in range(1, 7):
        col_letter = get_column_letter(col)
        if col_letter in ["D", "E", "F"]:
            ws.column_dimensions[col_letter].width = 21
        elif col_letter == "A":
            ws.column_dimensions[col_letter].width = 20
        else:
            max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1))
            ws.column_dimensions[col_letter].width = max(max_len + 2, 14)
    
    # Fill unstyled cells
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if cell.fill is None or cell.fill == PatternFill():
                cell.fill = white_fill


def create_assets_liabilities_sheet(wb, gl_df: pd.DataFrame, fund_name: str, 
                                   report_date: datetime, currency: str = "ETH"):
    """Create Assets and Liabilities sheet"""
    ws = wb.create_sheet(title="Assets and Liabilities")
    number_format_choice = get_number_format(currency)
    
    # Header formatting
    for row in range(1, 5):
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.fill = color_fill
            cell.font = font_subtitle if row > 2 else font_title

    ws["A1"] = fund_name
    ws["A1"].alignment = left
    ws["A1"].font = font_title

    ws["A2"] = "Statement of assets and liabilities (unaudited)"
    ws["A2"].alignment = left
    ws["A2"].font = font_title

    ws["A3"] = f"As of {report_date.strftime('%B %d, %Y')}"
    ws["A3"].alignment = left
    ws["A3"].font = Font(name=font_name, color="FFFFFF", size=12, bold=False)

    ws["A4"] = f"Reporting currency: {currency}"
    ws["A4"].alignment = left
    ws["A4"].font = Font(name=font_name, color="FFFFFF", size=12, bold=False)
    
    try:
        # Generate trial balance for the report date
        tb_df = generate_trial_balance_from_gl(gl_df, report_date)
        
        if tb_df.empty:
            ws["A6"] = "No trial balance data available"
            return
        
        # Apply bad debt netting logic (matching assets_liabilities.py)
        from .assets_liabilities import apply_bad_debt_netting
        
        # Filter out provision accounts for sister account netting
        provision_mask = tb_df['GL_Acct_Name'].str.contains('provision', case=False, na=False)
        provision_accounts = tb_df[provision_mask].copy() if provision_mask.any() else pd.DataFrame()
        
        # Remove provision accounts from main trial balance 
        tb_df = tb_df[~provision_mask]
        
        # Apply provision netting to loan receivables (sister account netting)
        tb_df = apply_bad_debt_netting(tb_df, provision_accounts)
        
        # Filter for assets and liabilities only
        assets_df = tb_df[tb_df['Category'] == 'Assets'].copy()
        liabilities_df = tb_df[tb_df['Category'] == 'Liabilities'].copy()
        
        # Remove zero balances
        assets_df = assets_df[abs(assets_df['Balance']) > 0.000001]
        liabilities_df = liabilities_df[abs(liabilities_df['Balance']) > 0.000001]
        
        # Sort by account name
        assets_df = assets_df.sort_values('GL_Acct_Name')
        liabilities_df = liabilities_df.sort_values('GL_Acct_Name')
        
        row = 6
        
        # Assets section
        def write_section(title, df, total_label, absolute_values=False):
            nonlocal row
            ws.cell(row=row, column=1, value=title).font = font_black
            ws.cell(row=row, column=1).alignment = left
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = light_color_fill
            row += 1
            
            section_total = 0.0
            for idx, (_, row_data) in enumerate(df.iterrows()):
                balance = abs(row_data['Balance']) if absolute_values else row_data['Balance']
                if abs(balance) <= 0.000001:
                    continue
                    
                fill = alt_fill if idx % 2 == 1 else no_fill
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = fill
                
                ws.cell(row=row, column=1, value=row_data['GL_Acct_Name']).font = font_black
                ws.cell(row=row, column=1).alignment = left
                
                cell = ws.cell(row=row, column=4, value=round(balance, 6))
                cell.font = font_black
                cell.alignment = right
                cell.number_format = number_format_choice
                
                section_total += balance
                row += 1
            
            # Total row
            ws.cell(row=row, column=1, value=total_label).font = font_black
            ws.cell(row=row, column=1).alignment = left
            cell = ws.cell(row=row, column=4, value=round(section_total, 6))
            cell.font = font_black
            cell.alignment = right
            cell.number_format = number_format_choice
            cell.border = Border(top=Side(style="thin", color="000000"))
            row += 2
            
            return section_total
        
        assets_total = write_section("Assets", assets_df, "Total Assets")
        liabilities_total = write_section("Liabilities", liabilities_df, "Total Liabilities", absolute_values=False)
        
        # Partners' Capital (Net Assets)
        net_assets = assets_total + liabilities_total  # liabilities_total is negative
        ws.cell(row=row, column=1, value="Partners' Capital").font = font_black_bold
        ws.cell(row=row, column=1).alignment = left
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = light_color_fill
        
        cell = ws.cell(row=row, column=4, value=round(net_assets, 6))
        cell.font = font_black_bold
        cell.alignment = right
        cell.number_format = number_format_choice
        row += 1
        
        # Footer fill
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = color_fill
        
    except Exception as e:
        ws["A6"] = f"Error generating assets & liabilities: {str(e)}"
    
    # Auto-adjust column widths
    for col in range(1, 5):
        col_letter = get_column_letter(col)
        if col_letter == "D":
            ws.column_dimensions[col_letter].width = 21
        else:
            max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1))
            ws.column_dimensions[col_letter].width = max(max_len + 2, 14)
    
    # Fill unstyled cells
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if cell.fill is None or cell.fill == PatternFill():
                cell.fill = white_fill


def create_operating_expenses_sheet(wb, gl_df: pd.DataFrame, fund_name: str, 
                                   report_date: datetime, currency: str = "ETH"):
    """Create Operating Expense Schedule sheet"""
    ws = wb.create_sheet(title="Operating Expense Schedule")
    number_format_choice = get_number_format(currency)
    
    # Calculate comparison date (previous month end)
    from datetime import timedelta
    comp_date = report_date.replace(day=1) - timedelta(days=1)
    
    # Create header section with extended range for this sheet
    for row in range(1, 10):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.fill = color_fill
            cell.font = font_subtitle if row > 2 else font_title
    
    ws["A1"] = fund_name
    ws["A1"].alignment = left
    ws["A1"].font = font_title
    
    ws["A2"] = "Schedule of operating expenses (unaudited)"
    ws["A2"].alignment = left
    ws["A2"].font = font_title
    
    ws["A3"] = f"For the period {(comp_date + timedelta(days=1)).strftime('%m/%d/%Y')} to {report_date.strftime('%m/%d/%Y')}"
    ws["A3"].alignment = left
    ws["A3"].font = Font(name=font_name, color="FFFFFF", size=12, bold=False)
    
    ws["A4"] = f"Reporting currency: {currency}"
    ws["A4"].alignment = left
    ws["A4"].font = Font(name=font_name, color="FFFFFF", size=12, bold=False)
    
    ws["A6"] = "Operating expense detail"
    ws["A6"].font = Font(name=font_name, color="0047BA", bold=True)
    ws["A6"].alignment = left
    
    try:
        # Generate trial balances for period comparison
        current_tb = generate_trial_balance_from_gl(gl_df, report_date)
        previous_tb = generate_trial_balance_from_gl(gl_df, comp_date)
        
        if current_tb.empty:
            ws["A8"] = "No expense data available"
            return
        
        # Get expense accounts only
        expense_accounts = current_tb[current_tb['Category'] == 'Expenses'].copy()
        
        if expense_accounts.empty:
            ws["A8"] = "No expense accounts found"
            return
        
        # Merge with previous balances
        if not previous_tb.empty:
            expense_data = pd.merge(
                expense_accounts[['GL_Acct_Number', 'GL_Acct_Name', 'Category', 'Balance']],
                previous_tb[['GL_Acct_Number', 'Balance']],
                on='GL_Acct_Number',
                how='left',
                suffixes=('_current', '_previous')
            )
        else:
            expense_data = expense_accounts.copy()
            expense_data['Balance_previous'] = 0
            expense_data['Balance_current'] = expense_data['Balance']
        
        # Fill missing values and calculate changes
        expense_data['Balance_previous'] = expense_data['Balance_previous'].fillna(0)
        expense_data['Balance_current'] = expense_data['Balance_current'].fillna(0)
        expense_data['Change'] = expense_data['Balance_current'] - expense_data['Balance_previous']
        
        # Calculate paid amount (begin + accrual - end)
        expense_data['Paid'] = expense_data['Balance_previous'] + expense_data['Change'] - expense_data['Balance_current']
        
        # Filter out zero rows
        expense_data = expense_data[
            (abs(expense_data['Balance_current']) > 0.000001) |
            (abs(expense_data['Balance_previous']) > 0.000001) |
            (abs(expense_data['Change']) > 0.000001) |
            (abs(expense_data['Paid']) > 0.000001)
        ]
        
        # Column headers
        headers = [
            "Account category", "GL number", "Accrual description",
            "Begin balances", "Accrual for the period", "Paid during the period",
            "Ending balances", "YTD accrual", "YTD paid"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col_num, value=header)
            cell.font = font_black
            cell.fill = light_color_fill
            cell.border = thin_border
            cell.alignment = left if col_num in [1, 3] else right
        
        # Data rows
        row = 8
        totals = [0.0] * 6  # For columns 4-9
        
        for idx, (_, expense_row) in enumerate(expense_data.iterrows()):
            fill = alt_fill if idx % 2 == 1 else no_fill
            
            # Prepare values
            begin_val = expense_row['Balance_previous']
            accrual_val = expense_row['Change']
            paid_val = expense_row['Paid']
            end_val = expense_row['Balance_current']
            
            def format_value(val):
                return round(val, 6) if abs(val) > 0.000001 else 0.0
            
            values = [
                expense_row['Category'],
                str(int(expense_row['GL_Acct_Number'])),
                expense_row['GL_Acct_Name'],
                format_value(begin_val),
                format_value(accrual_val), 
                format_value(paid_val),
                format_value(end_val),
                format_value(end_val),  # YTD accrual = ending balance
                format_value(paid_val)   # YTD paid = paid amount
            ]
            
            # Update totals
            for i, val in enumerate(values[3:]):
                if isinstance(val, (int, float)):
                    totals[i] += val
            
            # Write row
            for col_num, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_num, value=val)
                cell.font = font_black
                cell.fill = fill
                cell.alignment = left if col_num in [1, 3] else right
                
                if col_num >= 4 and isinstance(val, (int, float)):
                    cell.number_format = number_format_choice
            
            row += 1
        
        # Total row
        total_values = ["Total", "", ""] + [round(val, 6) for val in totals]
        for col_num, val in enumerate(total_values, 1):
            cell = ws.cell(row=row, column=col_num, value=val)
            cell.font = Font(name=font_name, bold=True, color="000000")
            cell.fill = light_color_fill
            cell.alignment = left if col_num in [1, 3] else right
            
            if col_num >= 4 and isinstance(val, (int, float)):
                cell.number_format = number_format_choice
        
    except Exception as e:
        ws["A8"] = f"Error generating operating expenses: {str(e)}"
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        col_letter = get_column_letter(col)
        if col_letter == "A":
            ws.column_dimensions[col_letter].width = 20
        elif col == 3:  # Accrual description
            ws.column_dimensions[col_letter].width = 50
        else:
            max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(7, ws.max_row + 1))
            ws.column_dimensions[col_letter].width = max(max_len + 2, 12)
    
    # Fill unstyled cells
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if cell.fill is None or cell.fill == PatternFill():
                cell.fill = white_fill


def create_management_fee_sheet(wb, gl_df: pd.DataFrame, fund_name: str, 
                               report_date: datetime, currency: str = "ETH"):
    """Create Management Fee calculation sheet"""
    ws = wb.create_sheet(title="Management Fee")
    number_format_choice = get_number_format(currency)
    
    # Calculate comparison date (previous month end)
    from datetime import timedelta
    comp_date = report_date.replace(day=1) - timedelta(days=1)
    
    # Header rows with styles
    for row in range(1, 5):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.fill = color_fill
            cell.font = font_subtitle if row > 2 else font_title
    
    ws["A1"] = fund_name
    ws["A1"].alignment = left
    ws["A1"].font = font_title
    
    ws["A2"] = "Management fee calculation (unaudited)"
    ws["A2"].alignment = left
    ws["A2"].font = font_title
    
    ws["A3"] = f"For the period {(comp_date + timedelta(days=1)).strftime('%m/%d/%Y')} to {report_date.strftime('%m/%d/%Y')}"
    ws["A3"].alignment = left
    ws["A3"].font = Font(name=font_name, color="FFFFFF", size=12, bold=False)
    
    ws["A4"] = f"Reporting currency: {currency}"
    ws["A4"].alignment = left
    ws["A4"].font = Font(name=font_name, color="FFFFFF", size=12, bold=False)
    
    try:
        # Check if GL data has required columns for management fee lookup
        if 'account_name' not in gl_df.columns:
            ws["A6"] = "No valid GL data available for management fee calculation"
            return
        
        # Load LP commitments data
        lp_df = load_LP_commitments_file()
        
        if lp_df.empty:
            ws["A6"] = "No LP commitments data available"
            return
        
        # Filter management fee expenses from GL 
        mgmt_fee_gl = gl_df[gl_df['account_name'].str.contains('management.*fee', case=False, na=False)]
        
        if mgmt_fee_gl.empty:
            ws["A6"] = "No management fee expense entries found"
            return
        
        # Column headers
        headers = [
            "Investor no", "Investor name", "Class", 
            "Management fee base", "Management fee rate %", "Management fee"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_num, value=header)
            cell.font = font_black
            cell.fill = light_color_fill
            cell.border = thin_border
            cell.alignment = right if col_num >= 4 else left
        
        # Process LP data
        management_fee_data = []
        
        for _, lp_row in lp_df.iterrows():
            try:
                lp_id = lp_row.get('reporting_limited_partner_ID', lp_row.get('limited_partner_ID', 'N/A'))
                investor_name = lp_row.get('lp_name', 'Unknown LP')
                share_class = lp_row.get('share_class', 'Class A')
                
                # Get commitment amount and management fee rate
                commitment = float(lp_row.get('commitment_amount', 0))
                mgmt_rate = float(lp_row.get('mgt_fee_1_rate', 0))
                mgmt_rate_display = f"{mgmt_rate * 100:.2f}%"
                
                # Calculate management fee from GL entries
                if currency == 'ETH':
                    total_fee = pd.to_numeric(mgmt_fee_gl['debit_crypto'], errors='coerce').fillna(0).sum() if not mgmt_fee_gl.empty else 0
                else:
                    total_fee = pd.to_numeric(mgmt_fee_gl['debit_USD'], errors='coerce').fillna(0).sum() if 'debit_USD' in mgmt_fee_gl.columns and not mgmt_fee_gl.empty else 0
                
                # For simplicity, divide equally among LPs
                lp_count = len(lp_df)
                lp_fee = total_fee / lp_count if lp_count > 0 else 0
                
                management_fee_data.append([
                    lp_id,
                    investor_name,
                    share_class,
                    round(commitment, 6),
                    mgmt_rate_display,
                    round(lp_fee, 6)
                ])
                
            except Exception as e:
                print(f"Warning: Error processing LP row: {e}")
                continue
        
        # If no valid data, show placeholder
        if not management_fee_data:
            management_fee_data = [[
                "N/A",
                "No LP Data Available",
                "N/A", 
                0.0,
                "0.00%",
                0.0
            ]]
        
        # Insert data rows
        row = 7
        for data_row in management_fee_data:
            for col_num, val in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col_num, value=val)
                cell.font = font_black
                cell.fill = no_fill
                cell.alignment = right if col_num >= 4 else left
                
                if col_num in [4, 6] and isinstance(val, (int, float)):
                    cell.number_format = number_format_choice
            row += 1
        
        # Total row
        total_base_sum = sum(row[3] for row in management_fee_data if isinstance(row[3], (int, float)))
        total_fee_sum = sum(row[5] for row in management_fee_data if isinstance(row[5], (int, float)))
        
        ws.cell(row=row, column=2, value="Total:").font = Font(name=font_name, bold=True)
        ws.cell(row=row, column=2).alignment = left
        
        ws.cell(row=row, column=4, value=round(total_base_sum, 6)).font = Font(name=font_name, bold=True)
        ws.cell(row=row, column=4).alignment = right
        ws.cell(row=row, column=4).number_format = number_format_choice
        
        ws.cell(row=row, column=6, value=round(total_fee_sum, 6)).font = Font(name=font_name, bold=True)
        ws.cell(row=row, column=6).alignment = right
        ws.cell(row=row, column=6).number_format = number_format_choice
        
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = alt_fill
        
    except Exception as e:
        print(f"Error loading LP commitments data: {e}")
        ws["A6"] = f"Error loading LP data: {str(e)}"
    
    # Auto-adjust column widths
    for col in range(1, 7):
        col_letter = get_column_letter(col)
        if col_letter == "A":
            ws.column_dimensions[col_letter].width = 13
        else:
            max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(6, ws.max_row + 1))
            ws.column_dimensions[col_letter].width = max(max_len + 2, 15)
    
    # Fill unstyled cells
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if cell.fill is None or cell.fill == PatternFill():
                cell.fill = white_fill


def generate_excel_report(gl_df: pd.DataFrame, fund_name: str, report_date: datetime, 
                         currency: str = "ETH", selected_fund: str = None) -> BytesIO:
    """
    Generate complete Excel report with ALL modules together
    
    Args:
        gl_df: General Ledger DataFrame
        fund_name: Name of the fund
        report_date: Reporting date
        currency: Reporting currency (ETH or USD)
        selected_fund: Fund identifier for specific calculations
        
    Returns:
        BytesIO object containing the Excel file
    """
    try:
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        print(f"DEBUG - Generating Excel report for {fund_name} on {report_date}")
        print(f"DEBUG - GL data shape: {gl_df.shape}")
        print(f"DEBUG - Creating ALL sheets: Account Statement (with NAV Changes), Trial Balance, Assets & Liabilities, Operating Expenses, Management Fee")
        
        # Create ALL sheets (all modules together)
        create_account_statement_sheet(wb, gl_df, fund_name, report_date, currency)
        create_trial_balance_sheet(wb, gl_df, fund_name, report_date, currency)
        create_assets_liabilities_sheet(wb, gl_df, fund_name, report_date, currency)
        create_operating_expenses_sheet(wb, gl_df, fund_name, report_date, currency)
        create_management_fee_sheet(wb, gl_df, fund_name, report_date, currency)
        
        print(f"DEBUG - Created {len(wb.worksheets)} worksheets: {[ws.title for ws in wb.worksheets]}")
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Get content and check size without affecting the BytesIO state
        content = output.getvalue()
        print(f"DEBUG - Excel generation: Content type: {type(content)}")
        print(f"DEBUG - Excel generation: Content size: {len(content)} bytes")
        print(f"DEBUG - Excel generation: Content is bytes: {isinstance(content, bytes)}")
        
        # Return a new BytesIO with the content to ensure clean state
        final_output = BytesIO(content)
        final_output.seek(0)
        
        print(f"DEBUG - Excel generation: final_output type: {type(final_output)}")
        print(f"DEBUG - Excel generation: final_output has getvalue: {hasattr(final_output, 'getvalue')}")
        
        # Verify the final output before returning
        test_content = final_output.getvalue()
        final_output.seek(0)  # Reset position after test
        print(f"DEBUG - Excel generation: final_output getvalue() returns: {type(test_content)}, size: {len(test_content)}")
        
        return final_output
        
    except Exception as e:
        print(f"ERROR - Excel generation failed: {e}")
        import traceback
        traceback.print_exc()
        
        # Return minimal workbook with error message
        wb = Workbook()
        ws = wb.active
        ws.title = "Error"
        ws["A1"] = f"Error generating report: {str(e)}"
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output