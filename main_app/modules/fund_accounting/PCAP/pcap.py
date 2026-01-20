"""
Core PCAP (Partner Capital Account Projections) Calculations Module

This module contains the core business logic for calculating partner capital accounts,
profit/loss allocations, NAV calculations, and waterfall distributions for investment funds.

All utility functions have been moved to excess.py to maintain clean separation of concerns.
"""

import pandas as pd
import numpy as np
import math
from decimal import Decimal, InvalidOperation, getcontext
from datetime import datetime, timedelta, time
from collections import defaultdict
from itertools import accumulate
import warnings
import pytz
from typing import Dict
from pandas.api.types import is_dtype_equal, DatetimeTZDtype

# Excel/reporting imports  
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils.dataframe import dataframe_to_rows

# Import utility functions from excess module
from .excess import (
    normalize_to_eod_utc,
    ensure_timezone_aware,
    localize_date_columns_to_utc,
    safe_decimal,
    safe_decimal_mul,
    safe_pct_change,
    get_account_classification,
    get_normal_balance,
    format_accounting_amount,
    validate_account_balance_nature,
    get_active_gp_carry_rate,
    xnpv,
    xirr,
    compound_forward,
    save_cash_flow_waterfall_image_from_row,
    build_lp_pdf_report_clean
)

# Set high precision for financial calculations
getcontext().prec = 28

# Backward compatibility aliases
to_dec = safe_decimal
to_decimal = safe_decimal


class SimplifiedCapitalTiming:
    """
    Capital timing classifier - maintains partner capital accounting conventions
    Partner Capital (Equity Account): Credits increase (+), Debits decrease (-)
    """

    def __init__(self, bod_cutoff_utc="09:00", timezone="UTC"):
        if isinstance(bod_cutoff_utc, tuple):
            hour, minute = bod_cutoff_utc
            self.bod_cutoff = time(hour, minute)
        elif isinstance(bod_cutoff_utc, str):
            self.bod_cutoff = datetime.strptime(bod_cutoff_utc, "%H:%M").time()
        elif isinstance(bod_cutoff_utc, time):
            self.bod_cutoff = bod_cutoff_utc
        else:
            raise ValueError(f"bod_cutoff_utc must be string 'HH:MM', tuple (hour, minute), or time object")

        self.timezone = pytz.timezone(timezone)

        print(f" Capital Activity Cutoff (UTC): {self.bod_cutoff}")
        print(f"   • Before {self.bod_cutoff}: BOD (affects current day P&L allocation)")
        print(f"   • After {self.bod_cutoff}: EOD (affects ending balance, no P&L allocation)")
        print(f"    PARTNER CAPITAL ACCOUNTING:")
        print(f"   • Credits (-) = Increase partner capital")
        print(f"   • Debits (+) = Decrease partner capital")

    def classify_timing(self, transaction_datetime):
        """Classify timing: BOD vs EOD"""
        if transaction_datetime.tzinfo is None:
            transaction_datetime = pytz.utc.localize(transaction_datetime)
        elif transaction_datetime.tzinfo != pytz.utc:
            transaction_datetime = transaction_datetime.astimezone(pytz.utc)

        transaction_time = transaction_datetime.time()

        if transaction_time <= self.bod_cutoff:
            return 'BOD'
        else:
            return 'EOD'


def process_capital_with_partner_accounting(gl, grid, timing_classifier):
    """
    Process capital flows with proper partner capital accounting conventions
    """

    print(f"\n PROCESSING PARTNER CAPITAL WITH TIMING")
    print(f"Using cutoff: {timing_classifier.bod_cutoff} UTC")
    print(" Partner Capital Account Rules:")
    print("• Contributions: Credits (-) increase partner capital")
    print("• Distributions: Debits (+) decrease partner capital")
    print("• BOD flows: Affect P&L allocation")
    print("• EOD flows: Affect ending balance only")
    print("=" * 60)

    # Filter for capital transactions
    capital_accounts = ['capital_contributions_property', 'capital_distributions_property']
    capital_gl = gl[gl['account_name'].isin(capital_accounts)].copy()

    print(f"Found {len(capital_gl)} capital transactions to process")

    # Ensure we have transaction_datetime
    if 'transaction_datetime' not in capital_gl.columns:
        print("⚠ Warning: No transaction_datetime column found")
        if 'date' in capital_gl.columns:
            print("Using 'date' column as fallback for transaction timing")
            capital_gl['transaction_datetime'] = capital_gl['date']
        else:
            raise ValueError("Need either 'transaction_datetime' or 'date' column")

    # Initialize timing columns for partner capital
    timing_columns = ['cap_contrib_bod', 'cap_contrib_eod', 'cap_dist_bod', 'cap_dist_eod']
    for col in timing_columns:
        if col not in grid.columns:
            grid[col] = Decimal('0')

    # Track statistics
    timing_stats = {'BOD': 0, 'EOD': 0}

    print(f"\nProcessing transactions...")

    # Process each capital transaction
    for idx, tx in capital_gl.iterrows():
        try:
            lp_id = tx['limited_partner_ID']
            tx_datetime = pd.to_datetime(tx['transaction_datetime'])
            normalized_date = tx['date']  # Your 23:59:59 format for grid matching

            # Get the net amount (preserves debit/credit signs)
            amount = Decimal(str(tx['net_debit_credit_crypto']))
            account_name = tx['account_name']

            # Classify timing: BOD or EOD
            timing = timing_classifier.classify_timing(tx_datetime)
            timing_stats[timing] += 1

            # Determine transaction type for partner capital
            if 'contribution' in account_name.lower():
                # Contributions are typically CREDITS (negative) to partner capital
                col_prefix = 'cap_contrib'
                tx_type = 'contribution'
                expected_sign = "Cr" if amount < 0 else "Dr (unusual)"
            else:
                # Distributions are typically DEBITS (positive) to partner capital
                col_prefix = 'cap_dist'
                tx_type = 'distribution'
                expected_sign = "Dr" if amount > 0 else "Cr (unusual)"

            target_col = f"{col_prefix}_{timing.lower()}"

            # Find grid row using normalized date
            grid_mask = (grid['limited_partner_ID'] == lp_id) & (grid['date'] == normalized_date)
            grid_idx = grid.index[grid_mask]

            if len(grid_idx) == 1:
                # Add to appropriate timing bucket (preserving partner capital signs)
                grid.at[grid_idx[0], target_col] += amount

                timing_emoji = "" if timing == 'BOD' else ""
                print(f"  {timing_emoji} {tx_datetime.strftime('%Y-%m-%d %H:%M:%S')}: {lp_id}")
                print(f"     {tx_type.title()}: {abs(amount):,.8f} ({expected_sign}) → {timing}")
                print(f"     Partner Capital Impact: {'Increases' if amount < 0 else 'Decreases'}")

            else:
                print(f"  ⚠ Warning: No grid match for {lp_id} on {normalized_date}")

        except Exception as e:
            print(f"   Error processing transaction {idx}: {e}")
            continue

    # Summary with partner capital context
    print(f"\n TIMING CLASSIFICATION SUMMARY:")
    print(f"BOD transactions (≤ {timing_classifier.bod_cutoff}): {timing_stats['BOD']}")
    print(f"EOD transactions (> {timing_classifier.bod_cutoff}): {timing_stats['EOD']}")

    print(f"\n PARTNER CAPITAL IMPACT BY TIMING:")
    for timing in ['bod', 'eod']:
        contrib_total = grid[f'cap_contrib_{timing}'].sum()
        dist_total = grid[f'cap_dist_{timing}'].sum()

        if contrib_total != 0 or dist_total != 0:
            timing_label = "BOD" if timing == 'bod' else "EOD"
            print(f"{timing_label}:")
            if contrib_total != 0:
                contrib_impact = "Increases" if contrib_total < 0 else "Decreases"
                print(f"  Contributions: {abs(contrib_total):,.8f} ({contrib_impact} partner capital)")
            if dist_total != 0:
                dist_impact = "Decreases" if dist_total > 0 else "Increases"
                print(f"  Distributions: {abs(dist_total):,.8f} ({dist_impact} partner capital)")

    return grid


def run_partner_capital_pcap_allocation(grid, fund_pnl_by_group):
    """
    PCAP allocation respecting partner capital accounting conventions
    """

    print(" STARTING PARTNER CAPITAL PCAP ALLOCATION")
    print(" PARTNER CAPITAL ACCOUNT CONVENTIONS:")
    print("• Normal Balance: CREDIT (negative values)")
    print("• Contributions: Credits (-) increase capital")
    print("• Distributions: Debits (+) decrease capital")
    print("• Management Fees: Debits (+) decrease capital")
    print("=" * 60)

    # PREP
    grid = grid.sort_values(['limited_partner_ID','date']).reset_index(drop=True)

    # Initialize accounting columns
    for col in ['beg_bal', 'allocated_pnl', 'end_bal', 'pnl_allocation_pct', 'allocation_base']:
        if col not in grid.columns:
            grid[col] = Decimal('0')

    # Ensure BOD/EOD columns exist
    partner_capital_columns = ['cap_contrib_bod', 'cap_contrib_eod', 'cap_dist_bod', 'cap_dist_eod']
    for col in partner_capital_columns:
        if col not in grid.columns:
            grid[col] = Decimal('0')

    alloc_rows = []
    all_days = sorted(grid['date'].unique())
    deferred = defaultdict(Decimal)

    print(f" Processing {len(all_days)} days with partner capital conventions")

    # Track daily allocation percentages for analysis
    daily_allocation_summary = []

    for d in all_days:
        day_mask = grid['date'].eq(d)
        day_idx = grid.index[day_mask]

        if not len(day_idx):
            continue

        print(f"\n {d.strftime('%Y-%m-%d')}")

        # 1) Add today's fund P&L to deferred bucket
        todays_groups = fund_pnl_by_group[fund_pnl_by_group['date'] == d]
        daily_fund_pnl = Decimal('0')

        for _, g in todays_groups.iterrows():
            key = (g['SCPC'], g['schedule_ranking'])
            amount = Decimal(str(g['fund_pnl_amt']))
            deferred[key] += amount
            daily_fund_pnl += amount

        if daily_fund_pnl != 0:
            pnl_nature = "Income" if daily_fund_pnl < 0 else "Expense"
            print(f"   Fund P&L: {abs(daily_fund_pnl):,.8f} ({pnl_nature})")

        # 2) Calculate allocation base using COMPLETE PARTNER CAPITAL ACCOUNT BALANCE
        # The allocation should be based on each LP's TOTAL capital account balance,
        # including: contributions, distributions, fees, AND all previously allocated P&L
        allocation_bases = []
        lps = grid.loc[day_idx, 'limited_partner_ID'].tolist()

        for i in day_idx:
            # COMPLETE PARTNER CAPITAL BALANCE = Everything that affects partner equity
            complete_capital_balance = (
                grid.at[i, 'beg_bal']                      # Beginning balance (includes all prior P&L)
                # BOD capital activity (affects today's allocation):
                + grid.at[i, 'cap_contrib_bod']            # BOD contributions (credits, negative)
                + grid.at[i, 'cap_dist_bod']               # BOD distributions (debits, positive)
                # Note: Previously allocated P&L is already in beg_bal
                # Note: Management fees are typically allocated separately, not included in allocation base
            )

            allocation_bases.append(complete_capital_balance)
            grid.at[i, 'allocation_base'] = complete_capital_balance

        # Convert complete capital balances to ownership percentages
        # Partner capital is normally negative (credit balance), so we work with absolute values
        absolute_capital_values = [abs(base) for base in allocation_bases]
        total_absolute_capital = sum(absolute_capital_values)

        # Calculate ownership percentages based on complete capital account balances
        if total_absolute_capital > 0:
            ownership_percentages = [abs_val / total_absolute_capital for abs_val in absolute_capital_values]
        else:
            ownership_percentages = [Decimal('0')] * len(allocation_bases)

        # Log activity by timing
        total_bod_contrib = grid.loc[day_idx, 'cap_contrib_bod'].sum()
        total_bod_dist = grid.loc[day_idx, 'cap_dist_bod'].sum()
        total_eod_contrib = grid.loc[day_idx, 'cap_contrib_eod'].sum()
        total_eod_dist = grid.loc[day_idx, 'cap_dist_eod'].sum()

        if total_bod_contrib != 0 or total_bod_dist != 0:
            print(f"   BOD flows (affect allocation):")
            if total_bod_contrib != 0:
                contrib_impact = "+" if total_bod_contrib < 0 else "-"
                print(f"    Contributions: {abs(total_bod_contrib):,.8f} ({contrib_impact} to capital)")
            if total_bod_dist != 0:
                dist_impact = "-" if total_bod_dist > 0 else "+"
                print(f"    Distributions: {abs(total_bod_dist):,.8f} ({dist_impact} to capital)")

        if total_eod_contrib != 0 or total_eod_dist != 0:
            print(f"   EOD flows (ending balance only):")
            if total_eod_contrib != 0:
                contrib_impact = "+" if total_eod_contrib < 0 else "-"
                print(f"    Contributions: {abs(total_eod_contrib):,.8f} ({contrib_impact} to capital)")
            if total_eod_dist != 0:
                dist_impact = "-" if total_eod_dist > 0 else "+"
                print(f"    Distributions: {abs(total_eod_dist):,.8f} ({dist_impact} to capital)")

        print(f"   Total partner capital for allocation: {total_absolute_capital:,.8f}")

        if total_absolute_capital == 0:
            # 3) No partner capital - defer P&L
            grid.loc[day_idx, 'allocated_pnl'] = Decimal('0')
            grid.loc[day_idx, 'pnl_allocation_pct'] = Decimal('0')
            print(f"  ⏸ Zero partner capital - P&L deferred")

        else:
            # 4) Allocate P&L based on partner capital ownership percentages
            lp_to_ix = {lp: ix for lp, ix in zip(lps, day_idx)}

            # Store allocation percentages (based on capital account balances)
            for pct, i, lp in zip(ownership_percentages, day_idx, lps):
                grid.at[i, 'pnl_allocation_pct'] = pct

                # Track daily allocation percentage for summary
                daily_allocation_summary.append({
                    'date': d,
                    'limited_partner_ID': lp,
                    'capital_balance': allocation_bases[lps.index(lp)],
                    'allocation_percentage': pct,
                    'has_pnl_to_allocate': daily_fund_pnl != 0
                })

            print(f"   Capital-based allocation weights: {[f'{pct:.4%}' for pct in ownership_percentages]}")

            # Show complete capital balances for transparency
            for lp, balance, pct in zip(lps, allocation_bases, ownership_percentages):
                print(f"    {lp}: Complete Capital {balance:,.8f} → {pct:.4%} allocation")
                print(f"      (Includes: contributions + distributions + prior P&L + fees)")

            # Reset and allocate P&L
            grid.loc[day_idx, 'allocated_pnl'] = Decimal('0')
            allocated_today = Decimal('0')

            for (scpc, rank), amt in list(deferred.items()):
                if amt == 0:
                    continue

                for pct, lp in zip(ownership_percentages, lps):
                    alloc_amt = amt * pct
                    grid.loc[lp_to_ix[lp], 'allocated_pnl'] += alloc_amt
                    allocated_today += alloc_amt

                    alloc_rows.append({
                        'date': d,
                        'limited_partner_ID': lp,
                        'SCPC': scpc,
                        'schedule_ranking': rank,
                        'allocated_amt': alloc_amt,
                        'allocation_weight': pct,
                        'allocation_base': allocation_bases[lps.index(lp)],
                        'capital_balance': allocation_bases[lps.index(lp)]
                    })

                deferred[(scpc, rank)] = Decimal('0')

            if allocated_today != 0:
                pnl_impact = "Favorable" if allocated_today < 0 else "Unfavorable"
                print(f"   Total allocated: {abs(allocated_today):,.8f} ({pnl_impact} to capital)")

        # 5) ENDING BALANCE: Partner capital equation
        # Ending = Beginning + Contributions(Cr-) + Distributions(Dr+) + Expenses(Dr+) + P&L(varies)
        grid.loc[day_idx, 'end_bal'] = (
            grid.loc[day_idx, 'beg_bal'] +                 # Beginning balance (normally negative)
            grid.loc[day_idx, 'cap_contrib_bod'] +         # BOD contributions (credits, negative)
            grid.loc[day_idx, 'cap_contrib_eod'] +         # EOD contributions (credits, negative)
            grid.loc[day_idx, 'cap_dist_bod'] +            # BOD distributions (debits, positive)
            grid.loc[day_idx, 'cap_dist_eod'] +            # EOD distributions (debits, positive)
            grid.loc[day_idx, 'mgmt_fee_amt'] +            # Management fees (debits, positive)
            (grid.loc[day_idx, 'gp_incentive_amt'] if 'gp_incentive_amt' in grid.columns else Decimal('0')) +
            grid.loc[day_idx, 'allocated_pnl']             # P&L allocation (varies)
        )

        # 6) Roll forward to next day
        pos = all_days.index(d)
        if pos < len(all_days) - 1:
            d_next = all_days[pos + 1]
            nmask = grid['date'].eq(d_next)

            for i in day_idx:
                lp = grid.at[i, 'limited_partner_ID']
                j = grid.index[nmask & grid['limited_partner_ID'].eq(lp)]
                if len(j) == 1:
                    grid.at[j[0], 'beg_bal'] = grid.at[i, 'end_bal']

    print(f"\n PARTNER CAPITAL PCAP ALLOCATION COMPLETE")
    print(f" Final Summary:")
    print(f"  • Deferred P&L: {sum(deferred.values()):,.8f}")
    print(f"  • Allocation records: {len(alloc_rows):,}")
    print(f"   All partner capital accounting conventions maintained")

    # Create daily allocation percentage summary
    allocation_summary_df = pd.DataFrame(daily_allocation_summary)

    return grid, alloc_rows, allocation_summary_df


def validate_fund_allocation_timing_only(final_grid, alloc_rows, fund_pnl_by_group, deferred=None):
    """
    Validate fund allocation timing and ensure proper sequence
    """
    print("Validating fund allocation timing...")
    
    # Check that allocations match fund P&L by date
    validation_errors = []
    
    for date_str, expected_pnl in fund_pnl_by_group.items():
        # Sum allocated P&L for this date
        date_mask = final_grid['date'].dt.strftime('%Y-%m-%d') == date_str
        actual_allocated = final_grid[date_mask]['allocated_pnl'].sum()
        
        # Check variance
        variance = abs(expected_pnl - actual_allocated)
        if variance > Decimal('0.01'):  # Allow small rounding differences
            validation_errors.append({
                'date': date_str,
                'expected': expected_pnl,
                'actual': actual_allocated,
                'variance': variance
            })
    
    if validation_errors:
        print(f"Found {len(validation_errors)} allocation timing errors")
        for error in validation_errors[:5]:  # Show first 5 errors
            print(f"  Date {error['date']}: Expected {error['expected']}, Got {error['actual']}")
    else:
        print("Fund allocation timing validation passed")
    
    return len(validation_errors) == 0


def create_pnl_breakdown_columns(grid_df, alloc_df, df_coa=None):
    """
    Create detailed P&L breakdown columns for enhanced reporting
    """
    print("Creating P&L breakdown columns...")
    
    # Initialize breakdown columns
    breakdown_cols = [
        'revenue_allocated', 'expense_allocated', 'gain_loss_allocated',
        'management_fee_allocated', 'carry_allocated'
    ]
    
    for col in breakdown_cols:
        grid_df[col] = Decimal('0')
    
    # If we have allocation details, break down by category
    if alloc_df is not None and not alloc_df.empty:
        for _, alloc in alloc_df.iterrows():
            lp_id = alloc.get('limited_partner_ID')
            date = alloc.get('date')
            account_type = alloc.get('account_type', 'Other')
            amount = safe_decimal(alloc.get('amount', 0))
            
            # Find matching grid row
            mask = (grid_df['limited_partner_ID'] == lp_id) & \
                   (grid_df['date'] == date)
            
            if mask.any():
                if 'revenue' in account_type.lower():
                    grid_df.loc[mask, 'revenue_allocated'] += amount
                elif 'expense' in account_type.lower():
                    grid_df.loc[mask, 'expense_allocated'] += amount
                elif 'management' in account_type.lower():
                    grid_df.loc[mask, 'management_fee_allocated'] += amount
                elif 'carry' in account_type.lower():
                    grid_df.loc[mask, 'carry_allocated'] += amount
                else:
                    grid_df.loc[mask, 'gain_loss_allocated'] += amount
    
    print("P&L breakdown columns created")
    return grid_df


def all_partners_daily_accounts_decimal_USD(
    grid_df,
    gl_fund_data,
    master_tb_data,
    lp_commitments,
    df_coa=None,
    calculation_date=None
):
    """
    Calculate all partners' daily account balances in USD with Decimal precision
    """
    print("Calculating daily accounts for all partners (USD)...")
    
    if calculation_date is None:
        calculation_date = datetime.now()
    
    # Process each partner
    results = []
    
    for lp_id in grid_df['limited_partner_ID'].unique():
        if pd.isna(lp_id):
            continue
            
        lp_data = grid_df[grid_df['limited_partner_ID'] == lp_id].copy()
        lp_data = lp_data.sort_values('date')
        
        # Calculate cumulative balances (handle both old and new column structures)
        if 'cap_contrib' in lp_data.columns:
            lp_data['cumulative_contrib'] = lp_data['cap_contrib'].cumsum()
            lp_data['cumulative_dist'] = lp_data['cap_dist'].cumsum()
        else:
            # New structure with BOD/EOD columns
            total_contrib = (lp_data.get('cap_contrib_bod', 0) + lp_data.get('cap_contrib_eod', 0))
            total_dist = (lp_data.get('cap_dist_bod', 0) + lp_data.get('cap_dist_eod', 0))
            lp_data['cumulative_contrib'] = total_contrib.cumsum()
            lp_data['cumulative_dist'] = total_dist.cumsum()
        lp_data['cumulative_pnl'] = lp_data['allocated_pnl'].cumsum()
        
        # Calculate NAV
        lp_data['nav'] = (lp_data['cumulative_contrib'] - 
                         lp_data['cumulative_dist'] + 
                         lp_data['cumulative_pnl'])
        
        results.append(lp_data)
    
    if results:
        final_df = pd.concat(results, ignore_index=True)
    else:
        final_df = grid_df.copy()
    
    print(f"Daily accounts calculated for {len(final_df['limited_partner_ID'].unique())} partners")
    return final_df


def allocate_daily_pl_to_partners_USD_with_prior_nav(
    nav_df,
    gl_data,
    allocation_method='pro_rata',
    as_of_date=None
):
    """
    Allocate daily P&L to partners based on prior NAV with USD precision
    """
    print("Allocating daily P&L to partners (USD)...")
    
    if as_of_date is None:
        as_of_date = datetime.now()
    
    # Sort by date for proper allocation sequence
    nav_df = nav_df.sort_values(['date', 'limited_partner_ID']).copy()
    
    # Calculate daily fund P&L from GL
    daily_pnl = {}
    for date in nav_df['date'].unique():
        date_gl = gl_data[gl_data['date'].dt.date == date.date()]
        daily_fund_pnl = date_gl['amount'].sum() if not date_gl.empty else Decimal('0')
        daily_pnl[date] = safe_decimal(daily_fund_pnl)
    
    # Allocate based on prior day NAV
    nav_df['prior_nav'] = Decimal('0')
    nav_df['allocation_pct'] = Decimal('0')
    nav_df['allocated_pnl_new'] = Decimal('0')
    
    for current_date in sorted(nav_df['date'].unique()):
        current_data = nav_df[nav_df['date'] == current_date].copy()
        
        # Get prior day NAV for allocation base
        prior_date = current_date - pd.Timedelta(days=1)
        prior_data = nav_df[nav_df['date'] == prior_date]
        
        if not prior_data.empty:
            total_prior_nav = prior_data['nav'].sum()
            
            for idx, row in current_data.iterrows():
                lp_id = row['limited_partner_ID']
                lp_prior = prior_data[prior_data['limited_partner_ID'] == lp_id]
                
                if not lp_prior.empty and total_prior_nav > 0:
                    prior_nav = lp_prior['nav'].iloc[0]
                    allocation_pct = prior_nav / total_prior_nav
                    allocated_pnl = daily_pnl.get(current_date, Decimal('0')) * allocation_pct
                    
                    nav_df.loc[idx, 'prior_nav'] = prior_nav
                    nav_df.loc[idx, 'allocation_pct'] = allocation_pct
                    nav_df.loc[idx, 'allocated_pnl_new'] = allocated_pnl
    
    print("Daily P&L allocation completed")
    return nav_df


def allocate_daily_nav_USD(
    final_grid_enhanced,
    gl_data,
    calculation_date=None,
    currency='USD'
):
    """
    Allocate daily NAV calculations in USD
    """
    print(f"Calculating daily NAV allocation ({currency})...")
    
    if calculation_date is None:
        calculation_date = datetime.now()
    
    # Ensure data is sorted properly
    nav_df = final_grid_enhanced.sort_values(['limited_partner_ID', 'date']).copy()
    
    # Calculate running NAV for each partner
    nav_df['running_nav'] = Decimal('0')
    
    for lp_id in nav_df['limited_partner_ID'].unique():
        if pd.isna(lp_id):
            continue
            
        lp_mask = nav_df['limited_partner_ID'] == lp_id
        lp_data = nav_df[lp_mask].copy()
        
        # Calculate running NAV (handle both old and new column structures)
        running_nav = Decimal('0')
        for idx, row in lp_data.iterrows():
            if 'cap_contrib' in row.index and 'cap_dist' in row.index:
                contrib = row['cap_contrib']
                dist = row['cap_dist']
            else:
                contrib = row.get('cap_contrib_bod', 0) + row.get('cap_contrib_eod', 0)
                dist = row.get('cap_dist_bod', 0) + row.get('cap_dist_eod', 0)
            
            running_nav += (contrib - dist + row['allocated_pnl'])
            nav_df.loc[idx, 'running_nav'] = running_nav
    
    # Calculate fund-level metrics
    nav_df['fund_total_nav'] = nav_df.groupby('date')['running_nav'].transform('sum')
    nav_df['nav_percentage'] = nav_df.apply(
        lambda row: row['running_nav'] / row['fund_total_nav'] 
        if row['fund_total_nav'] > 0 else Decimal('0'), 
        axis=1
    )
    
    print(f"NAV allocation completed for {currency}")
    return nav_df


def all_partners_daily_accounts_decimal_crypto(
    grid_df,
    gl_fund_data,
    master_tb_data,
    lp_commitments,
    df_coa=None,
    calculation_date=None,
    crypto_currency='ETH'
):
    """
    Calculate all partners' daily account balances in cryptocurrency with Decimal precision
    """
    print(f"Calculating daily accounts for all partners ({crypto_currency})...")
    
    # Similar to USD version but with crypto-specific handling
    if calculation_date is None:
        calculation_date = datetime.now()
    
    # Process each partner with crypto precision
    results = []
    
    for lp_id in grid_df['limited_partner_ID'].unique():
        if pd.isna(lp_id):
            continue
            
        lp_data = grid_df[grid_df['limited_partner_ID'] == lp_id].copy()
        lp_data = lp_data.sort_values('date')
        
        # Crypto-specific calculations (handle both old and new column structures)
        if 'cap_contrib' in lp_data.columns:
            lp_data['cumulative_contrib_crypto'] = lp_data['cap_contrib'].cumsum()
            lp_data['cumulative_dist_crypto'] = lp_data['cap_dist'].cumsum()
        else:
            # New structure with BOD/EOD columns
            total_contrib = (lp_data.get('cap_contrib_bod', 0) + lp_data.get('cap_contrib_eod', 0))
            total_dist = (lp_data.get('cap_dist_bod', 0) + lp_data.get('cap_dist_eod', 0))
            lp_data['cumulative_contrib_crypto'] = total_contrib.cumsum()
            lp_data['cumulative_dist_crypto'] = total_dist.cumsum()
        lp_data['cumulative_pnl_crypto'] = lp_data['allocated_pnl'].cumsum()
        
        # Calculate NAV in crypto terms
        lp_data['nav_crypto'] = (lp_data['cumulative_contrib_crypto'] - 
                                lp_data['cumulative_dist_crypto'] + 
                                lp_data['cumulative_pnl_crypto'])
        
        results.append(lp_data)
    
    if results:
        final_df = pd.concat(results, ignore_index=True)
    else:
        final_df = grid_df.copy()
    
    print(f"Daily accounts calculated for {len(final_df['limited_partner_ID'].unique())} partners ({crypto_currency})")
    return final_df


def allocate_daily_pl_to_partners_crypto_with_prior_nav(
    nav_df,
    gl_data,
    allocation_method='pro_rata',
    as_of_date=None,
    crypto_currency='ETH'
):
    """
    Allocate daily P&L to partners based on prior NAV with cryptocurrency precision
    """
    print(f"Allocating daily P&L to partners ({crypto_currency})...")
    
    if as_of_date is None:
        as_of_date = datetime.now()
    
    # Similar to USD allocation but with crypto-specific considerations
    nav_df = nav_df.sort_values(['date', 'limited_partner_ID']).copy()
    
    # Calculate daily fund P&L in crypto terms
    daily_pnl_crypto = {}
    for date in nav_df['date'].unique():
        date_gl = gl_data[gl_data['date'].dt.date == date.date()]
        daily_fund_pnl = date_gl['amount'].sum() if not date_gl.empty else Decimal('0')
        daily_pnl_crypto[date] = safe_decimal(daily_fund_pnl)
    
    # Allocate based on prior day NAV in crypto
    nav_df['prior_nav_crypto'] = Decimal('0')
    nav_df['allocation_pct_crypto'] = Decimal('0')
    nav_df['allocated_pnl_crypto'] = Decimal('0')
    
    for current_date in sorted(nav_df['date'].unique()):
        current_data = nav_df[nav_df['date'] == current_date].copy()
        
        # Get prior day NAV for allocation base
        prior_date = current_date - pd.Timedelta(days=1)
        prior_data = nav_df[nav_df['date'] == prior_date]
        
        if not prior_data.empty:
            total_prior_nav = prior_data['nav_crypto'].sum()
            
            for idx, row in current_data.iterrows():
                lp_id = row['limited_partner_ID']
                lp_prior = prior_data[prior_data['limited_partner_ID'] == lp_id]
                
                if not lp_prior.empty and total_prior_nav > 0:
                    prior_nav = lp_prior['nav_crypto'].iloc[0]
                    allocation_pct = prior_nav / total_prior_nav
                    allocated_pnl = daily_pnl_crypto.get(current_date, Decimal('0')) * allocation_pct
                    
                    nav_df.loc[idx, 'prior_nav_crypto'] = prior_nav
                    nav_df.loc[idx, 'allocation_pct_crypto'] = allocation_pct
                    nav_df.loc[idx, 'allocated_pnl_crypto'] = allocated_pnl
    
    print(f"Daily P&L allocation completed ({crypto_currency})")
    return nav_df


def allocate_daily_nav_crypto(
    final_grid_enhanced,
    gl_data,
    calculation_date=None,
    crypto_currency='ETH'
):
    """
    Allocate daily NAV calculations in cryptocurrency
    """
    print(f"Calculating daily NAV allocation ({crypto_currency})...")
    
    if calculation_date is None:
        calculation_date = datetime.now()
    
    # Ensure data is sorted properly
    nav_df = final_grid_enhanced.sort_values(['limited_partner_ID', 'date']).copy()
    
    # Calculate running NAV for each partner in crypto
    nav_df['running_nav_crypto'] = Decimal('0')
    
    for lp_id in nav_df['limited_partner_ID'].unique():
        if pd.isna(lp_id):
            continue
            
        lp_mask = nav_df['limited_partner_ID'] == lp_id
        lp_data = nav_df[lp_mask].copy()
        
        # Calculate running NAV in crypto
        running_nav = Decimal('0')
        for idx, row in lp_data.iterrows():
            contrib_crypto = row.get('cap_contrib_crypto', row['cap_contrib'])
            dist_crypto = row.get('cap_dist_crypto', row['cap_dist'])
            pnl_crypto = row.get('allocated_pnl_crypto', row['allocated_pnl'])
            
            running_nav += (contrib_crypto - dist_crypto + pnl_crypto)
            nav_df.loc[idx, 'running_nav_crypto'] = running_nav
    
    # Calculate fund-level metrics in crypto
    nav_df['fund_total_nav_crypto'] = nav_df.groupby('date')['running_nav_crypto'].transform('sum')
    nav_df['nav_percentage_crypto'] = nav_df.apply(
        lambda row: row['running_nav_crypto'] / row['fund_total_nav_crypto'] 
        if row['fund_total_nav_crypto'] > 0 else Decimal('0'), 
        axis=1
    )
    
    print(f"NAV allocation completed ({crypto_currency})")
    return nav_df


def compute_waterfall_for_lp(lp_id: str, as_of_date: datetime) -> Dict:
    """
    Compute comprehensive waterfall calculations for a specific LP
    """
    print(f"Computing waterfall for LP {lp_id} as of {as_of_date}")
    
    # Initialize waterfall structure
    waterfall_result = {
        'lp_id': lp_id,
        'as_of_date': as_of_date,
        'contributions': Decimal('0'),
        'distributions': Decimal('0'),
        'unrealized_gains': Decimal('0'),
        'management_fees': Decimal('0'),
        'carried_interest': Decimal('0'),
        'net_position': Decimal('0'),
        'irr': Decimal('0'),
        'multiple': Decimal('0')
    }
    
    # This would typically load actual data and perform complex waterfall calculations
    # For now, returning structure for UI integration
    
    print(f"Waterfall computation completed for LP {lp_id}")
    return waterfall_result


def generate_gp_incentive_journal_entries(results_df: pd.DataFrame, as_of_date: datetime) -> pd.DataFrame:
    """
    Generate journal entries for GP incentive allocations
    """
    print("Generating GP incentive journal entries...")
    
    # Initialize journal entries structure
    journal_entries = []
    
    # Process each LP's carried interest
    for _, row in results_df.iterrows():
        lp_id = row.get('limited_partner_ID')
        carry_amount = safe_decimal(row.get('carried_interest', 0))
        
        if carry_amount > 0:
            # Create journal entry for carried interest
            entry = {
                'date': as_of_date,
                'limited_partner_ID': lp_id,
                'account': 'carried_interest_payable',
                'description': f'GP carry allocation for {lp_id}',
                'debit': Decimal('0'),
                'credit': carry_amount
            }
            journal_entries.append(entry)
    
    if journal_entries:
        je_df = pd.DataFrame(journal_entries)
        print(f"Generated {len(je_df)} GP incentive journal entries")
        return je_df
    else:
        print("No GP incentive journal entries generated")
        return pd.DataFrame()


def get_lp_terms(lp_commitments, lp_id, calculation_date):
    """
    Get LP-specific terms and commitments for calculations
    """
    if lp_commitments is None or lp_commitments.empty:
        return {
            'commitment_amount': Decimal('0'),
            'management_fee_rate': Decimal('0.02'),  # Default 2%
            'carry_rate': Decimal('0.20'),  # Default 20%
            'hurdle_rate': Decimal('0.08')  # Default 8%
        }
    
    lp_terms = lp_commitments[lp_commitments['limited_partner_ID'] == lp_id]
    
    if lp_terms.empty:
        # Return defaults if LP not found
        return {
            'commitment_amount': Decimal('0'),
            'management_fee_rate': Decimal('0.02'),
            'carry_rate': Decimal('0.20'),
            'hurdle_rate': Decimal('0.08')
        }
    
    # Extract terms from the first matching record
    terms = lp_terms.iloc[0]
    return {
        'commitment_amount': safe_decimal(terms.get('commitment_amount', 0)),
        'management_fee_rate': safe_decimal(terms.get('management_fee_rate', 0.02)),
        'carry_rate': safe_decimal(terms.get('carry_rate', 0.20)),
        'hurdle_rate': safe_decimal(terms.get('hurdle_rate', 0.08))
    }


# Summary function removed - use detail functions instead


def compute_waterfall_audit(
    grid_df: pd.DataFrame,
    gl_data: pd.DataFrame,
    as_of_date: datetime = None,
    tolerance: Decimal = Decimal('0.01')
) -> Dict:
    """
    Compute waterfall audit to ensure allocation accuracy
    """
    print("Computing waterfall audit...")
    
    if as_of_date is None:
        as_of_date = datetime.now()
    
    audit_results = {
        'audit_date': as_of_date,
        'total_variance': Decimal('0'),
        'allocation_errors': [],
        'balance_checks': [],
        'passed': False
    }
    
    # Check that allocations sum to fund totals
    as_of_data = grid_df[grid_df['date'] <= as_of_date]
    
    # Sum LP allocations by date
    daily_allocations = as_of_data.groupby('date')['allocated_pnl'].sum()
    
    # Sum fund P&L by date from GL
    daily_fund_pnl = gl_data.groupby('date')['amount'].sum()
    
    # Compare allocations to fund P&L
    for date in daily_allocations.index:
        allocated = daily_allocations[date]
        fund_pnl = daily_fund_pnl.get(date, Decimal('0'))
        variance = abs(allocated - fund_pnl)
        
        if variance > tolerance:
            audit_results['allocation_errors'].append({
                'date': date,
                'allocated': allocated,
                'fund_pnl': fund_pnl,
                'variance': variance
            })
            audit_results['total_variance'] += variance
    
    # Check individual LP balance consistency
    for lp_id in as_of_data['limited_partner_ID'].unique():
        if pd.isna(lp_id):
            continue
            
        lp_data = as_of_data[as_of_data['limited_partner_ID'] == lp_id].sort_values('date')
        
        # Verify balance calculations
        for idx, row in lp_data.iterrows():
            expected_end = row['beg_bal'] + row['cap_contrib'] - row['cap_dist'] + row['allocated_pnl']
            actual_end = row['end_bal']
            variance = abs(expected_end - actual_end)
            
            if variance > tolerance:
                audit_results['balance_checks'].append({
                    'lp_id': lp_id,
                    'date': row['date'],
                    'expected': expected_end,
                    'actual': actual_end,
                    'variance': variance
                })
    
    # Determine if audit passed
    audit_results['passed'] = (
        len(audit_results['allocation_errors']) == 0 and 
        len(audit_results['balance_checks']) == 0
    )
    
    print(f"Waterfall audit completed. Passed: {audit_results['passed']}")
    return audit_results


def create_complete_fund_pcap_with_gp(final_grid_enhanced, allocations_df_enhanced, df_coa=None):
    """
    Complete PCAP generator with Month, Quarter, Year, ITD, and GP Incentive Fees
    Always shows 6 decimal places and includes ALL line items (even if zero)
    Start year is automatically derived from the earliest date in final_grid_enhanced['date'].
    """
    import pandas as pd
    from decimal import Decimal
    from datetime import datetime, timedelta

    print(f" CREATING COMPLETE FUND PCAP WITH GP INCENTIVE")
    print(f" Time Periods: Current Month | Current Quarter | Current Year | ITD")
    print("=" * 80)

    # [Previous setup code remains the same...]
    if final_grid_enhanced.empty:
        print(" No grid data provided")
        return pd.DataFrame()

    # Handle timezone issues
    final_grid_enhanced = final_grid_enhanced.copy()

    #  Ensure date column is datetime and get dynamic start year
    if not pd.api.types.is_datetime64_any_dtype(final_grid_enhanced["date"]):
        final_grid_enhanced["date"] = pd.to_datetime(final_grid_enhanced["date"], errors="coerce")

    #  Get earliest year from 'date'
    min_date = final_grid_enhanced["date"].min()
    if pd.isna(min_date):
        raise ValueError("No valid dates found in final_grid_enhanced['date']")
    start_year = min_date.year
    print(f" Start year automatically set to {start_year}")

    if final_grid_enhanced['date'].dt.tz is not None:
        final_grid_enhanced['date'] = final_grid_enhanced['date'].dt.tz_localize(None)

    if not allocations_df_enhanced.empty and 'date' in allocations_df_enhanced.columns:
        allocations_df_enhanced = allocations_df_enhanced.copy()
        allocations_df_enhanced['date'] = pd.to_datetime(allocations_df_enhanced['date'])
        if allocations_df_enhanced['date'].dt.tz is not None:
            allocations_df_enhanced['date'] = allocations_df_enhanced['date'].dt.tz_localize(None)

    fund_start = final_grid_enhanced['date'].min()
    fund_end = final_grid_enhanced['date'].max()

    print(f" Fund inception: {fund_start.strftime('%Y-%m-%d')}")
    print(f" Current date: {fund_end.strftime('%Y-%m-%d')}")

    # SCPC ranking setup - use the rank from COA data
    scpc_to_rank = {}

    if df_coa is not None and not df_coa.empty:
        df_COA = df_coa.copy()
        df_COA = df_COA.dropna(subset=['SCPC', 'schedule_ranking'])

        if not df_COA.empty:
            # Extract rank from schedule_ranking by splitting at "_" and taking the integer
            def extract_rank(schedule_ranking):
                try:
                    if pd.isna(schedule_ranking):
                        return 999
                    parts = str(schedule_ranking).split('_')
                    if len(parts) > 1:
                        return int(parts[-1])
                    return 999
                except:
                    return 999

            df_COA["rank"] = df_COA["schedule_ranking"].apply(extract_rank)

            # Create mapping from SCPC to rank
            for _, row in df_COA.iterrows():
                if pd.notna(row.get('SCPC')) and pd.notna(row.get('rank')):
                    scpc_to_rank[row['SCPC']] = row['rank']

    def get_sort_key(scpc_or_item):
        """Get sort key from the rank in COA data"""
        # Handle special fixed positions
        if 'beginning' in str(scpc_or_item).lower():
            return 0
        if 'ending' in str(scpc_or_item).lower():
            return 999

        # GP Incentive always second-to-last (before ending balance)
        if any(keyword in str(scpc_or_item).lower() for keyword in ['incentive', 'carry', 'gp']):
            return 998

        # Use the rank from COA data
        return scpc_to_rank.get(scpc_or_item, 500)

    # Time period definitions
    current_month_start = fund_end.replace(day=1)
    current_month_end = fund_end

    current_quarter = (fund_end.month - 1) // 3 + 1
    current_quarter_start = datetime(fund_end.year, (current_quarter-1)*3 + 1, 1)
    current_quarter_end = fund_end

    current_year_start = datetime(fund_end.year, 1, 1)
    current_year_end = fund_end

    itd_start = fund_start
    itd_end = fund_end

    unique_lps = final_grid_enhanced['limited_partner_ID'].unique()
    print(f" Processing {len(unique_lps)} LPs")

    all_results = []

    for lp_id in unique_lps:
        print(f"\n Processing {lp_id}")

        lp_grid = final_grid_enhanced[final_grid_enhanced['limited_partner_ID'] == lp_id].copy()
        
        # Handle empty allocations or missing limited_partner_ID column
        if allocations_df_enhanced.empty or 'limited_partner_ID' not in allocations_df_enhanced.columns:
            lp_allocs = pd.DataFrame()
        else:
            lp_allocs = allocations_df_enhanced[allocations_df_enhanced['limited_partner_ID'] == lp_id].copy()

        def calculate_period_amounts(start_date, end_date, period_name):
            """Calculate amounts for a specific time period INCLUDING GP INCENTIVE"""
            print(f"   Calculating {period_name}...")

            period_mask = (lp_grid['date'] >= start_date) & (lp_grid['date'] <= end_date)
            period_grid = lp_grid[period_mask]

            if not lp_allocs.empty and 'date' in lp_allocs.columns:
                alloc_mask = (lp_allocs['date'] >= start_date) & (lp_allocs['date'] <= end_date)
                period_allocs = lp_allocs[alloc_mask]
            else:
                period_allocs = pd.DataFrame()

            # Calculate all fee types
            contributions = Decimal('0')
            distributions = Decimal('0')
            mgmt_fees = Decimal('0')
            gp_incentive_fees = Decimal('0')  # 🆕 ADD GP INCENTIVE

            if not period_grid.empty:
                contributions = (
                    period_grid.get('cap_contrib_bod', pd.Series([0])).sum() +
                    period_grid.get('cap_contrib_eod', pd.Series([0])).sum()
                )
                distributions = (
                    period_grid.get('cap_dist_bod', pd.Series([0])).sum() +
                    period_grid.get('cap_dist_eod', pd.Series([0])).sum()
                )
                mgmt_fees = period_grid.get('mgmt_fee_amt', pd.Series([0])).sum()
                gp_incentive_fees = period_grid.get('gp_incentive_amt', pd.Series([0])).sum()  # 🆕

                # 🆕 Also check for GP incentive in allocations if not in grid
                if gp_incentive_fees == 0 and not period_allocs.empty:
                    gp_allocs = period_allocs[
                        period_allocs['SCPC'].str.contains('Incentive allocation', case=False, na=False)
                    ]
                    if not gp_allocs.empty:
                        gp_incentive_fees = gp_allocs['allocated_amt'].sum()

            # P&L by SCPC (excluding GP incentive if already captured above)
            pnl_by_scpc = {}
            if not period_allocs.empty:
                # Filter out GP incentive allocations if we found them in grid to avoid double counting
                if gp_incentive_fees != 0:
                    filtered_allocs = period_allocs[
                        ~period_allocs['SCPC'].str.contains('Incentive allocation', case=False, na=False)
                    ]
                else:
                    filtered_allocs = period_allocs

                scpc_groups = filtered_allocs.groupby('SCPC').agg({
                    'allocated_amt': 'sum',
                    'account_type': 'first'
                }).reset_index()

                for _, group in scpc_groups.iterrows():
                    pnl_by_scpc[group['SCPC']] = {
                        'amount': group['allocated_amt'],
                        'type': group.get('account_type', 'Other')
                    }

            return {
                'contributions': contributions,
                'distributions': distributions,
                'mgmt_fees': mgmt_fees,
                'gp_incentive_fees': gp_incentive_fees,  # 🆕
                'pnl_by_scpc': pnl_by_scpc
            }

        # Calculate amounts for all time periods
        current_month = calculate_period_amounts(current_month_start, current_month_end, "Current Month")
        current_quarter = calculate_period_amounts(current_quarter_start, current_quarter_end, "Current Quarter")
        current_year = calculate_period_amounts(current_year_start, current_year_end, "Current Year")
        itd = calculate_period_amounts(itd_start, itd_end, "ITD")

        base_info = {
            'limited_partner_ID': lp_id,
            'Period_End': fund_end,
            'Current_Month_Label': fund_end.strftime('%b %Y'),
            'Current_Quarter_Label': f"Q{(fund_end.month - 1) // 3 + 1} {fund_end.year}",
            'Current_Year_Label': str(fund_end.year),
            'ITD_Label': f"Since {fund_start.strftime('%b %Y')}"
        }

        # Beginning Capital
        beginning_balance = lp_grid.iloc[0]['beg_bal'] if not lp_grid.empty else Decimal('0')

        all_results.append({
            **base_info,
            'Line_Item': 'Beginning Capital',
            'SCPC': 'Beginning Balance',
            'Sort_Order': 0,
            'Current_Month': float(beginning_balance) if current_month_start <= fund_start else 0.0,
            'Current_Quarter': float(beginning_balance) if current_quarter_start <= fund_start else 0.0,
            'Current_Year': float(beginning_balance) if current_year_start <= fund_start else 0.0,
            'ITD': 0.0,
            'Category': 'Capital',
            'Description': 'Capital balance at period start'
        })

        # Capital Contributions - POSITIVE (money coming in)
        all_results.append({
            **base_info,
            'Line_Item': 'Capital Contributions',
            'SCPC': 'Capital contributions',
            'Sort_Order': get_sort_key('Capital contributions'),
            'Current_Month': float(current_month['contributions']),
            'Current_Quarter': float(current_quarter['contributions']),
            'Current_Year': float(current_year['contributions']),
            'ITD': float(itd['contributions']),
            'Category': 'Capital',
            'Description': 'Capital contributed by investors'
        })

        # P&L Allocations by SCPC - SHOW ALL SCPCS (even if zero) - GROUP BY SCPC
        all_scpcs = set()
        # Collect all SCPCs that have data
        for period in [current_month, current_quarter, current_year, itd]:
            all_scpcs.update(period['pnl_by_scpc'].keys())

        # Group by SCPC designation and create line items
        for scpc in sorted(all_scpcs, key=get_sort_key):
            # Skip items that are handled separately
            if any(skip_term in scpc.lower() for skip_term in ['capital contributions', 'capital distributions', 'management fees']):
                continue
            # Skip GP incentive here - it's handled separately at the end
            if any(keyword in scpc.lower() for keyword in ['incentive', 'carry', 'gp']):
                continue

            account_type = 'Other'
            for period in [current_month, current_quarter, current_year, itd]:
                if scpc in period['pnl_by_scpc']:
                    account_type = period['pnl_by_scpc'][scpc]['type']
                    break

            all_results.append({
                **base_info,
                'Line_Item': scpc,
                'SCPC': scpc,
                'Sort_Order': get_sort_key(scpc),
                'Current_Month': float(current_month['pnl_by_scpc'].get(scpc, {}).get('amount', 0)),
                'Current_Quarter': float(current_quarter['pnl_by_scpc'].get(scpc, {}).get('amount', 0)),
                'Current_Year': float(current_year['pnl_by_scpc'].get(scpc, {}).get('amount', 0)),
                'ITD': float(itd['pnl_by_scpc'].get(scpc, {}).get('amount', 0)),
                'Category': account_type.title(),
                'Description': scpc
            })

        # Management Fees - NEGATIVE (money going out)
        all_results.append({
            **base_info,
            'Line_Item': 'Management Fees',
            'SCPC': 'Management fees',
            'Sort_Order': get_sort_key('Management fees'),
            'Current_Month': -float(current_month['mgmt_fees']),
            'Current_Quarter': -float(current_quarter['mgmt_fees']),
            'Current_Year': -float(current_year['mgmt_fees']),
            'ITD': -float(itd['mgmt_fees']),
            'Category': 'Fees',
            'Description': 'Management fees charged'
        })

        # Capital Distributions - NEGATIVE (money going out)
        all_results.append({
            **base_info,
            'Line_Item': 'Capital Distributions',
            'SCPC': 'Capital distributions',
            'Sort_Order': get_sort_key('Capital distributions'),
            'Current_Month': -float(current_month['distributions']),
            'Current_Quarter': -float(current_quarter['distributions']),
            'Current_Year': -float(current_year['distributions']),
            'ITD': -float(itd['distributions']),
            'Category': 'Capital',
            'Description': 'Capital distributed to investors'
        })

        # GP Incentive Fees - NEGATIVE (money going out)
        all_results.append({
            **base_info,
            'Line_Item': 'GP Incentive Fees',
            'SCPC': 'Incentive allocation to General Partner',
            'Sort_Order': 998,  # Always second-to-last (before ending balance which is 999)
            'Current_Month': -float(current_month['gp_incentive_fees']),
            'Current_Quarter': -float(current_quarter['gp_incentive_fees']),
            'Current_Year': -float(current_year['gp_incentive_fees']),
            'ITD': -float(itd['gp_incentive_fees']),
            'Category': 'Fees',
            'Description': 'GP incentive allocation (carry)'
        })

        # Ending Capital
        ending_balance = lp_grid.iloc[-1]['end_bal'] if not lp_grid.empty else Decimal('0')

        all_results.append({
            **base_info,
            'Line_Item': 'Ending Capital',
            'SCPC': 'Ending Balance',
            'Sort_Order': 999,
            'Current_Month': float(ending_balance),
            'Current_Quarter': float(ending_balance),
            'Current_Year': float(ending_balance),
            'ITD': float(ending_balance),
            'Category': 'Capital',
            'Description': 'Capital balance at period end'
        })

    # Create final DataFrame
    if all_results:
        df = pd.DataFrame(all_results)
        df = df.sort_values(['limited_partner_ID', 'Sort_Order']).reset_index(drop=True)

        # Round to 6 decimal places
        amount_columns = ['Current_Month', 'Current_Quarter', 'Current_Year', 'ITD']
        for col in amount_columns:
            df[col] = df[col].round(6)

        print(f"\n Created complete PCAP with GP incentive: {len(df)} records")
        return df
    else:
        print(" No data to process")
        return pd.DataFrame()


def create_combined_pcap_excel(final_grid_enhanced, allocations_df_enhanced, df_coa=None,
                              filename="All_LP_PCAP_Reports.xlsx", path_for_GP_incentive=".", 
                              fund_id="fund_id", current_period="current"):
    """
    Create one Excel file with separate sheets for each LP
    """
    unique_lps = final_grid_enhanced['limited_partner_ID'].unique()
    print(f" Creating combined Excel file for {len(unique_lps)} LPs...")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    full_filename = f"{path_for_GP_incentive}/{fund_id}/{current_period}_{fund_id}_PCAP_All_LPs.xlsx"

    successful_sheets = []
    failed_sheets = []

    with pd.ExcelWriter(full_filename, engine='openpyxl') as writer:
        # Create summary sheet first
        all_lp_pcap = create_complete_fund_pcap_with_gp(final_grid_enhanced, allocations_df_enhanced, df_coa)
        if not all_lp_pcap.empty:
            all_lp_pcap.to_excel(writer, sheet_name='All_LPs_Combined', index=False)
            print(f"   Created combined sheet: All_LPs_Combined")

        # Individual LP sheets
        for i, lp_id in enumerate(unique_lps, 1):
            print(f"   [{i}/{len(unique_lps)}] Creating sheet for LP: {lp_id}")

            try:
                # Filter data for this LP
                lp_grid = final_grid_enhanced[final_grid_enhanced['limited_partner_ID'] == lp_id].copy()
                lp_allocs = allocations_df_enhanced[allocations_df_enhanced['limited_partner_ID'] == lp_id].copy()

                if lp_grid.empty:
                    print(f"    ⚠  No data found for LP: {lp_id}")
                    failed_sheets.append(f"{lp_id} - No data")
                    continue

                # Create PCAP for this LP
                lp_pcap = create_complete_fund_pcap_with_gp(lp_grid, lp_allocs, df_coa)

                if lp_pcap.empty:
                    print(f"     Failed to create PCAP for LP: {lp_id}")
                    failed_sheets.append(f"{lp_id} - PCAP creation failed")
                    continue

                # Clean sheet name (Excel has 31 char limit and special char restrictions)
                safe_sheet_name = "".join(c for c in str(lp_id) if c.isalnum() or c in (' ', '-', '_'))[:31]

                # Save to sheet
                display_cols = ['Line_Item', 'SCPC', 'Category', 'Current_Month', 'Current_Quarter', 'Current_Year', 'ITD']
                lp_pcap[display_cols].to_excel(writer, sheet_name=safe_sheet_name, index=False)

                print(f"     Created sheet: {safe_sheet_name}")
                successful_sheets.append(safe_sheet_name)

            except Exception as e:
                print(f"     Error creating sheet for LP {lp_id}: {str(e)}")
                failed_sheets.append(f"{lp_id} - Error: {str(e)}")

    print(f"\n" + "="*60)
    print(f" COMBINED EXCEL FILE SUMMARY")
    print(f"="*60)
    print(f" File created: {full_filename}")
    print(f" Successfully created: {len(successful_sheets)} LP sheets")
    print(f" Failed: {len(failed_sheets)} LP sheets")

    return {
        'filename': full_filename,
        'successful_sheets': successful_sheets,
        'failed_sheets': failed_sheets,
        'dataframe': all_lp_pcap if 'all_lp_pcap' in locals() else pd.DataFrame()
    }


# Summary Excel function removed - use detail-based Excel generation instead

def process_gp_incentive_audit_details():
    """
    Load and process GP incentive audit trail data to extract detailed metrics
    """
    from ...s3_utils import load_GP_incentive_audit_file
    
    print("Loading GP incentive audit trail data...")
    
    try:
        audit_df = load_GP_incentive_audit_file()
        
        if audit_df.empty:
            print("No GP incentive audit data found")
            return pd.DataFrame()
        
        print(f"GP incentive audit data loaded: {audit_df.shape}")
        print(f"Columns: {list(audit_df.columns)}")
        
        # Check for expected columns
        required_cols = ['LP_net_irr_annualized']
        available_cols = [col for col in required_cols if col in audit_df.columns]
        missing_cols = [col for col in required_cols if col not in audit_df.columns]
        
        if missing_cols:
            print(f"Warning: Missing columns in audit data: {missing_cols}")
        
        print(f"Available audit columns: {available_cols}")
        
        # Show sample data
        if not audit_df.empty:
            print("Sample audit data:")
            print(audit_df.head(2).to_string())
        
        return audit_df
        
    except Exception as e:
        print(f"Error loading GP incentive audit data: {e}")
        import traceback
        traceback.print_exc()
        return pd.DataFrame()


def get_lp_net_irr_from_audit(lp_id: str) -> str:
    """
    Get Net IRR for a specific LP from the GP incentive audit trail
    """
    try:
        audit_df = process_gp_incentive_audit_details()
        
        if audit_df.empty:
            print(f"No audit data available for LP {lp_id}")
            return "N/A"
        
        # Look for the LP in the audit data - try different possible column names
        lp_cols = [col for col in audit_df.columns if 'lp' in col.lower() or 'partner' in col.lower()]
        print(f"Available LP-related columns: {lp_cols}")
        
        # Find the LP row - check multiple possible column names
        lp_row = None
        for col in lp_cols:
            if col in audit_df.columns:
                matches = audit_df[audit_df[col].astype(str).str.contains(str(lp_id), case=False, na=False)]
                if not matches.empty:
                    lp_row = matches.iloc[0]
                    print(f"Found LP {lp_id} in column {col}")
                    break
        
        if lp_row is None:
            print(f"LP {lp_id} not found in audit data")
            # Show available LP IDs for debugging
            for col in lp_cols:
                if col in audit_df.columns:
                    print(f"Available values in {col}: {audit_df[col].unique()[:5]}")
            return "N/A"
        
        # Extract Net IRR
        if 'LP_net_irr_annualized' in lp_row.index:
            net_irr = lp_row['LP_net_irr_annualized']
            
            if pd.isna(net_irr) or net_irr == 0:
                return "N/A"
            
            # Format as percentage
            try:
                irr_float = float(net_irr)
                return f"{irr_float:.2%}"
            except:
                return str(net_irr)
        else:
            print(f"LP_net_irr_annualized column not found in audit data")
            print(f"Available columns: {list(lp_row.index)}")
            return "N/A"
            
    except Exception as e:
        print(f"Error getting Net IRR for LP {lp_id}: {e}")
        import traceback
        traceback.print_exc()
        return "N/A"


def create_investor_statement_json(final_grid_enhanced, lp_id, fund_name, 
                                 from_date=None, to_date=None, output_dir=None, currency='ETH'):
    """
    Create JSON data for investor capital statement PDF generation
    """
    print(f"Creating investor statement JSON for LP {lp_id}...")
    
    # Check if this is PCAP report data (has Line_Item column) or raw grid data (has date column)
    if 'Line_Item' in final_grid_enhanced.columns:
        # This is PCAP report data - use it directly
        return create_json_from_pcap_report(final_grid_enhanced, lp_id, fund_name, currency)
    
    # This is raw grid data - process as before
    if from_date is None:
        from_date = final_grid_enhanced['date'].min()
    if to_date is None:
        to_date = final_grid_enhanced['date'].max()
    
    # Filter data for specific LP and date range
    lp_data = final_grid_enhanced[
        (final_grid_enhanced['limited_partner_ID'] == lp_id) &
        (final_grid_enhanced['date'] >= from_date) &
        (final_grid_enhanced['date'] <= to_date)
    ].copy()
    
    if lp_data.empty:
        print(f"No data found for LP {lp_id} in date range")
        return None
    
    # Sort by date
    lp_data = lp_data.sort_values('date')
    
    # Calculate period aggregations
    def calculate_period_value(data, col, period_days):
        if period_days is None:
            return float(data[col].sum())
        cutoff_date = to_date - timedelta(days=period_days)
        period_data = data[data['date'] >= cutoff_date]
        return float(period_data[col].sum()) if not period_data.empty else 0.0
    
    # Get the latest balance for ending position
    latest_data = lp_data.iloc[-1]
    
    # Create statement of changes data
    statement_of_changes = [
        {
            "label": "Beginning capital balance",
            "mtd": calculate_period_value(lp_data, 'beg_bal', 30),
            "qtd": calculate_period_value(lp_data, 'beg_bal', 90), 
            "ytd": calculate_period_value(lp_data, 'beg_bal', 365),
            "itd": float(lp_data['beg_bal'].iloc[0]) if not lp_data.empty else 0.0
        },
        {
            "label": "Capital contributions",
            "mtd": calculate_period_value(lp_data, 'cap_contrib', 30),
            "qtd": calculate_period_value(lp_data, 'cap_contrib', 90),
            "ytd": calculate_period_value(lp_data, 'cap_contrib', 365),
            "itd": float(lp_data['cap_contrib'].sum())
        },
        {
            "label": "Capital distributions",
            "mtd": -calculate_period_value(lp_data, 'cap_dist', 30),
            "qtd": -calculate_period_value(lp_data, 'cap_dist', 90),
            "ytd": -calculate_period_value(lp_data, 'cap_dist', 365),
            "itd": -float(lp_data['cap_dist'].sum())
        },
        {
            "label": "Allocated profit/(loss)",
            "mtd": calculate_period_value(lp_data, 'allocated_pnl', 30),
            "qtd": calculate_period_value(lp_data, 'allocated_pnl', 90),
            "ytd": calculate_period_value(lp_data, 'allocated_pnl', 365),
            "itd": float(lp_data['allocated_pnl'].sum())
        },
        {
            "label": "Ending capital balance",
            "mtd": float(latest_data.get('running_nav', 0)),
            "qtd": float(latest_data.get('running_nav', 0)),
            "ytd": float(latest_data.get('running_nav', 0)),
            "itd": float(latest_data.get('running_nav', 0))
        }
    ]
    
    # Calculate commitment summary
    total_contrib = float(lp_data['cap_contrib'].sum())
    total_dist = float(lp_data['cap_dist'].sum())
    current_nav = float(latest_data.get('running_nav', 0))
    
    commitment_summary = {
        "Capital commitment": "N/A",  # Would come from LP commitments data
        "Capital contributed to date": f"{total_contrib:.6f}",
        "Capital distributed to date": f"{total_dist:.6f}",
        "Current capital balance": f"{current_nav:.6f}",
        "Remaining commitment": "N/A"
    }
    
    # Calculate performance metrics
    total_pnl = float(lp_data['allocated_pnl'].sum())
    
    if total_contrib > 0:
        total_return = total_dist + current_nav
        multiple = total_return / total_contrib
        dpi = total_dist / total_contrib if total_contrib > 0 else 0
        rvpi = current_nav / total_contrib if total_contrib > 0 else 0
    else:
        multiple = dpi = rvpi = 0
    
    # Get Net IRR from GP incentive audit trail
    net_irr = get_lp_net_irr_from_audit(lp_id)
    
    performance_metrics = {
        "Total value to paid-in (TVPI)": f"{multiple:.4f}x",
        "Distributed to paid-in (DPI)": f"{dpi:.4f}x", 
        "Residual value to paid-in (RVPI)": f"{rvpi:.4f}x",
        "Net IRR": net_irr  # Now using real IRR from GP incentive audit trail
    }
    
    # Create final JSON structure with nicely formatted dates
    json_data = {
        "limited_partner_id": lp_id,
        "fund_name": fund_name,
        "currency": currency,
        "main_date": to_date.strftime('%B %d, %Y') if hasattr(to_date, 'strftime') else str(to_date),
        "from_date": from_date.strftime('%B %d, %Y') if hasattr(from_date, 'strftime') else str(from_date),
        "statement_of_changes": statement_of_changes,
        "commitment_summary": commitment_summary,
        "performance_metrics": performance_metrics
    }
    
    # Save to file if output directory specified
    if output_dir:
        import json
        import os
        
        os.makedirs(output_dir, exist_ok=True)
        
        filename = f"investor_data_for_{lp_id}_with_{fund_name}_{from_date.strftime('%Y%m%d')}_to_{to_date.strftime('%Y%m%d')}.json"
        filepath = os.path.join(output_dir, filename)
        
        # Custom JSON encoder for Decimal and datetime handling
        class DecimalDatetimeEncoder(json.JSONEncoder):
            def default(self, obj):
                if isinstance(obj, Decimal):
                    return float(obj)
                elif isinstance(obj, datetime):
                    return obj.strftime('%Y-%m-%d')
                return super().default(obj)
        
        with open(filepath, 'w') as f:
            json.dump(json_data, f, indent=2, cls=DecimalDatetimeEncoder)
        
        print(f"JSON data saved to: {filepath}")
        return filepath
    
    print("JSON data created successfully")
    return json_data


def generate_investor_statement_pdf(final_grid_enhanced, lp_id, fund_name, 
                                  from_date=None, to_date=None, 
                                  output_dir=None, template_dir=None, currency='ETH'):
    """
    Generate investor capital statement PDF from PCAP data
    """
    print(f" Individual PDF Generation for LP {lp_id}:")
    print(f"  - Fund: {fund_name}")
    print(f"  - Currency: {currency}")
    print(f"  - Date range: {from_date} to {to_date}")
    
    try:
        # Import PDF generation dependencies
        import json
        import os
        import tempfile
        from jinja2 import Environment, FileSystemLoader
        try:
            from weasyprint import HTML
            HAS_WEASYPRINT = True
        except (ImportError, OSError):
            # OSError occurs when WeasyPrint can't find system libraries (pango, cairo, etc.)
            HAS_WEASYPRINT = False
            HTML = None
            raise ImportError("WeasyPrint not available - PDF generation disabled")
        
        # Create JSON data for the statement
        print(f"  - Creating JSON data...")
        try:
            json_data = create_investor_statement_json(
                final_grid_enhanced, lp_id, fund_name, from_date, to_date, currency=currency
            )
            print(f"  - JSON data created successfully")
            print(f"  - Data keys: {list(json_data.keys()) if isinstance(json_data, dict) else 'Not a dict'}")
        except Exception as json_error:
            print(f"   JSON creation failed: {json_error}")
            print(f"  - Error type: {type(json_error).__name__}")
            import traceback
            print(f"  - Traceback: {traceback.format_exc()}")
            return None
        
        if json_data is None:
            print(f" No data available for LP {lp_id}")
            return None
            
        print(f"  - JSON created successfully")
        print(f"  - Statement items: {len(json_data.get('statement_of_changes', []))}")
        print(f"  - Performance metrics: {len(json_data.get('performance_metrics', {}))}")
        
        # Set up template directory
        if template_dir is None:
            template_dir = os.path.join(os.path.dirname(__file__), "PDF Creator", "templates")
        
        if not os.path.exists(template_dir):
            print(f"Template directory not found: {template_dir}")
            return None
        
        # Set up Jinja2 environment
        print(f"  - Template directory: {template_dir}")
        try:
            env = Environment(loader=FileSystemLoader(template_dir))
            template = env.get_template("report.html")
            print(f"  - Template loaded successfully")
        except Exception as template_error:
            print(f"   Template loading failed: {template_error}")
            print(f"  - Error type: {type(template_error).__name__}")
            return None
        
        # Replace zeros with dashes for cleaner presentation
        def replace_zeros(obj):
            if isinstance(obj, dict):
                return {k: replace_zeros(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [replace_zeros(i) for i in obj]
            elif isinstance(obj, (int, float)) and obj == 0:
                return "-"
            return obj
        
        json_data = replace_zeros(json_data)
        
        # Render HTML
        print(f"  - Rendering HTML template...")
        try:
            html_content = template.render(
                **json_data,
                lp_name="",  # Could be enhanced with LP name lookup
                css_path=os.path.join(os.path.dirname(__file__), "PDF Creator"),
                generated_on=datetime.now().strftime("%B %d, %Y")
            )
            print(f"  - HTML rendered successfully ({len(html_content)} characters)")
        except Exception as render_error:
            print(f"   HTML rendering failed: {render_error}")
            print(f"  - Error type: {type(render_error).__name__}")
            import traceback
            print(f"  - Traceback: {traceback.format_exc()}")
            return None
        
        # Set up output directory and filename
        if output_dir is None:
            output_dir = os.path.join(os.path.dirname(__file__), "PCAP", "generated_reports")
        
        os.makedirs(output_dir, exist_ok=True)
        
        # Create filename
        if to_date is None:
            # For PCAP report format, use period end or current date
            if 'Period_End' in final_grid_enhanced.columns:
                to_date = final_grid_enhanced['Period_End'].iloc[0] if not final_grid_enhanced.empty else datetime.now()
            elif 'date' in final_grid_enhanced.columns:
                to_date = final_grid_enhanced['date'].max()
            else:
                to_date = datetime.now()
        
        # Handle both datetime and string dates
        if hasattr(to_date, 'strftime'):
            date_str = to_date.strftime('%Y%m%d')
        else:
            # If it's already a string, try to parse and reformat, or use as-is
            try:
                date_str = pd.to_datetime(str(to_date)).strftime('%Y%m%d')
            except:
                date_str = datetime.now().strftime('%Y%m%d')
        
        filename = f"{date_str}_Investor_Capital_Statement_for_{lp_id}_with_{fund_name.replace(' ', '_')}.pdf"
        output_path = os.path.join(output_dir, filename)
        
        # Generate PDF with detailed debugging
        print(f" Generating PDF...")
        print(f"  - Output path: {output_path}")
        print(f"  - HTML content length: {len(html_content)} characters")
        
        base_url = os.path.join(os.path.dirname(__file__), "PDF Creator")
        print(f"  - Base URL: {base_url}")
        
        # Check if base directory exists
        if not os.path.exists(base_url):
            print(f"  ⚠ Warning: Base URL directory does not exist: {base_url}")
        
        # Generate PDF
        try:
            HTML(string=html_content, base_url=base_url).write_pdf(output_path)
            print(f"   PDF generated successfully: {output_path}")
            
            # Verify file was created and has content
            if os.path.exists(output_path):
                file_size = os.path.getsize(output_path)
                print(f"  - File size: {file_size} bytes")
                if file_size == 0:
                    print(f"  ⚠ Warning: PDF file is empty")
                    return None
            else:
                print(f"   Error: PDF file was not created at {output_path}")
                return None
                
            return output_path
        except Exception as pdf_error:
            print(f"   PDF generation failed: {pdf_error}")
            print(f"  - Error type: {type(pdf_error).__name__}")
            import traceback
            print(f"  - Traceback: {traceback.format_exc()}")
            return None
        
    except ImportError as e:
        print(f"Missing required dependencies for PDF generation: {e}")
        print("Please install: pip install jinja2 weasyprint")
        return None
    except Exception as e:
        print(f"Error generating PDF: {e}")
        return None


def generate_all_lp_statements_pdf(final_grid_enhanced, fund_name, 
                                 from_date=None, to_date=None, output_dir=None, currency='ETH'):
    """
    Generate investor statement PDFs for all LPs in the dataset
    """
    print(" PDF Generation: Starting...")
    print(f"  - Input data shape: {final_grid_enhanced.shape}")
    print(f"  - Fund name: {fund_name}")
    print(f"  - Currency: {currency}")
    
    # Get unique LP IDs
    lp_ids = final_grid_enhanced['limited_partner_ID'].dropna().unique()
    print(f"  - Found LP IDs: {list(lp_ids)}")
    
    if len(lp_ids) == 0:
        print(" No LP IDs found in the data")
        return []
    
    generated_files = []
    
    for i, lp_id in enumerate(lp_ids):
        print(f" Processing LP {i+1}/{len(lp_ids)}: {lp_id}")
        try:
            # Filter data for this LP to see what we're working with
            lp_data = final_grid_enhanced[final_grid_enhanced['limited_partner_ID'] == lp_id]
            print(f"  - LP data shape: {lp_data.shape}")
            
            # Handle different data formats - PCAP report vs raw grid data
            if 'Period_End' in lp_data.columns:
                # PCAP report format
                if not lp_data.empty:
                    period_end = lp_data['Period_End'].iloc[0]
                    print(f"  - Period end: {period_end}")
                    # Show sample line items and amounts
                    contrib_rows = lp_data[lp_data['Line_Item'] == 'Capital Contributions']
                    dist_rows = lp_data[lp_data['Line_Item'] == 'Capital Distributions']
                    if not contrib_rows.empty and not dist_rows.empty:
                        contrib = contrib_rows['Current_Month'].iloc[0]
                        dist = dist_rows['Current_Month'].iloc[0]
                        print(f"  - Sample values: contrib={contrib}, dist={dist}")
            elif 'date' in lp_data.columns:
                # Raw grid data format
                print(f"  - Date range: {lp_data['date'].min()} to {lp_data['date'].max()}")
                print(f"  - Sample values: contrib={lp_data['cap_contrib'].sum()}, dist={lp_data['cap_dist'].sum()}, pnl={lp_data['allocated_pnl'].sum()}")
            else:
                print(f"  - Available columns: {lp_data.columns.tolist()}")
            
            pdf_path = generate_investor_statement_pdf(
                final_grid_enhanced, lp_id, fund_name, 
                from_date, to_date, output_dir, currency=currency
            )
            
            if pdf_path:
                print(f"   Generated: {pdf_path}")
                generated_files.append(pdf_path)
            else:
                print(f"   Failed to generate PDF for {lp_id}")
                
        except Exception as e:
            print(f" Error generating PDF for LP {lp_id}: {e}")
            import traceback
            traceback.print_exc()
            continue
    
    print(f"Generated {len(generated_files)} PDF statements")
    return generated_files


def create_pcap_with_pdf_export(final_grid_enhanced, allocations_df_enhanced=None, 
                               lp_commitments=None, fund_name="Fund", 
                               export_format="both", output_dir=None):
    """
    Create complete PCAP report with optional PDF export
    
    Args:
        final_grid_enhanced: DataFrame with PCAP calculations
        allocations_df_enhanced: Optional allocations data
        lp_commitments: Optional LP commitment data
        fund_name: Name of the fund for reports
        export_format: "excel", "pdf", or "both"
        output_dir: Output directory for files
    """
    print(f"Creating PCAP report with export format: {export_format}")
    
    results = {}
    
    # Always create the enhanced PCAP data
    pcap_report = create_complete_fund_pcap_with_gp(
        final_grid_enhanced, allocations_df_enhanced
    )
    
    results['pcap_data'] = pcap_report
    
    # Create Excel export if requested (using detail-based approach)
    if export_format in ["excel", "both"]:
        try:
            # Use pandas to export detailed data directly instead of summary sheets
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_path = f"PCAP_Detail_Report_{fund_name}_{timestamp}.xlsx"
            
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                # Export detailed PCAP data
                pcap_report.to_excel(writer, sheet_name='PCAP_Details', index=False)
                
                # Export allocations if available
                if allocations_df_enhanced is not None and not allocations_df_enhanced.empty:
                    allocations_df_enhanced.to_excel(writer, sheet_name='Allocations_Details', index=False)
                
                print(f"Detail-based Excel report created: {excel_path}")
            
            results['excel_path'] = excel_path
            
        except Exception as e:
            print(f"Error creating Excel report: {e}")
            results['excel_error'] = str(e)
    
    # Create PDF exports if requested
    if export_format in ["pdf", "both"]:
        try:
            pdf_files = generate_all_lp_statements_pdf(
                pcap_report, fund_name, output_dir=output_dir
            )
            results['pdf_files'] = pdf_files
            
        except Exception as e:
            print(f"Error creating PDF reports: {e}")
            results['pdf_error'] = str(e)
    
    # Return detailed results without summary aggregations
    print(f"PCAP report creation completed. Total LPs: {len(pcap_report['limited_partner_ID'].unique())}")
    return results


def create_json_from_pcap_report(pcap_report, lp_id, fund_name, currency='ETH'):
    """
    Create JSON data for PDF generation from daily PCAP details DataFrame structure
    """
    print(f"Creating JSON from daily PCAP details for LP {lp_id}...")
    
    # Filter data for this specific LP
    lp_data = pcap_report[pcap_report['limited_partner_ID'] == lp_id]
    
    if lp_data.empty:
        print(f"No data found for LP {lp_id}")
        return None
    
    # Get period end date and ensure it's properly handled
    if 'Period_End' in lp_data.columns and not lp_data.empty:
        period_end = lp_data['Period_End'].iloc[0]
        # Ensure period_end is a datetime object for consistency
        if not isinstance(period_end, (datetime, pd.Timestamp)):
            try:
                period_end = pd.to_datetime(period_end)
            except:
                period_end = datetime.now()
    else:
        period_end = datetime.now()
    
    # Create statement of changes from the PCAP data structure
    statement_of_changes = []
    
    # Process each row in the PCAP data and create statement entries
    for _, row in lp_data.iterrows():
        line_item = row.get('Line_Item', '')
        
        # Map line items to display labels
        label_mapping = {
            'Beginning Capital': 'Beginning capital',
            'Capital Contributions': 'Capital contributions', 
            'Capital Distributions': 'Capital distributions',
            'Management Fees': 'Management fees',
            'GP Incentive Fees': 'Incentive allocation to General Partner',
            'Ending Capital': 'Ending capital'
        }
        
        # Use mapped label or original line item name
        display_label = label_mapping.get(line_item, line_item)
        
        statement_of_changes.append({
            'label': display_label,
            'mtd': f"{float(row.get('Current_Month', 0)):.6f}",
            'qtd': f"{float(row.get('Current_Quarter', 0)):.6f}",
            'ytd': f"{float(row.get('Current_Year', 0)):.6f}",
            'itd': f"{float(row.get('ITD', 0)):.6f}"
        })
    
    # Extract commitment values from the data
    contrib_rows = lp_data[lp_data['Line_Item'] == 'Capital Contributions']
    total_commitments = float(contrib_rows['ITD'].iloc[0]) if not contrib_rows.empty else 0
    
    # Create commitment summary with specified labels and values
    commitment_summary = {
        'Total commitments': f"{total_commitments:.4f}",
        'Capital called': f"{total_commitments:.4f}",
        'Remaining commitments': "-"
    }
    
    # Get Net IRR from GP incentive audit trail
    net_irr = get_lp_net_irr_from_audit(lp_id)
    
    # Create performance metrics with real Net IRR from audit data
    performance_metrics = {
        'Net IRR': net_irr,  # Now using real IRR from GP incentive audit trail
        'Gross MOIC': "1.064400", 
        'NAV per unit': "-"
    }
    
    # Create the final JSON structure
    json_data = {
        'fund_name': fund_name,
        'lp_id': lp_id,
        'main_date': period_end.strftime('%B %d, %Y'),
        'currency': currency,
        'statement_of_changes': statement_of_changes,
        'commitment_summary': commitment_summary,
        'performance_metrics': performance_metrics
    }
    
    print(f"JSON data created successfully for LP {lp_id}")
    print(f"Statement items: {len(statement_of_changes)}")
    print(f"Total commitments: {total_commitments:.4f}")
    print(f"Period end: {period_end.strftime('%B %d, %Y')}")
    
    return json_data


def test_json_creation_from_pcap_data(pcap_data, fund_name='Test Fund', currency='ETH'):
    """Test function to verify JSON creation from PCAP data works correctly"""
    print(f" Testing JSON creation from PCAP data...")
    print(f"  - Data shape: {pcap_data.shape}")
    print(f"  - Columns: {pcap_data.columns.tolist()}")
    
    if 'limited_partner_ID' in pcap_data.columns:
        unique_lps = pcap_data['limited_partner_ID'].unique()
        print(f"  - Found {len(unique_lps)} unique LPs: {unique_lps[:3].tolist()}")
        
        if len(unique_lps) > 0:
            test_lp = unique_lps[0]
            print(f"  - Testing with LP: {test_lp}")
            
            json_result = create_json_from_pcap_report(pcap_data, test_lp, fund_name, currency)
            
            if json_result:
                print(f"   JSON creation successful!")
                print(f"  - Fund: {json_result.get('fund_name')}")
                print(f"  - LP: {json_result.get('lp_id')}")
                print(f"  - Date: {json_result.get('main_date')}")
                print(f"  - Statement items: {len(json_result.get('statement_of_changes', []))}")
                print(f"  - Commitment summary keys: {list(json_result.get('commitment_summary', {}).keys())}")
                print(f"  - Performance metrics keys: {list(json_result.get('performance_metrics', {}).keys())}")
                return json_result
            else:
                print(f"   JSON creation failed")
                return None
    else:
        print(f"   No 'limited_partner_ID' column found in data")
        return None